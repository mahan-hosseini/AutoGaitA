# %% imports
from autogaita.gaita_res.utils import bin_num_to_percentages
from autogaita.group.group_utils import save_figures, write_issues_to_textfile
import os
import string
import pandas as pd
import numpy as np
from sklearn.preprocessing import StandardScaler
from sklearn.decomposition import PCA
import openpyxl
import seaborn as sns
import matplotlib.pyplot as plt
from matplotlib.animation import FuncAnimation, FFMpegWriter
import pdb

# %% constants
from autogaita.group.group_constants import (
    ID_COL,
    GROUP_COL,
    SC_PERCENTAGE_COL,
    PCA_BARPLOT_BARCOLOR,
    PCA_BARPLOT_LINECOLOR,
    PCA_CUSTOM_SCATTER_OUTER_SEPARATOR,
    PCA_CUSTOM_SCATTER_INNER_SEPARATOR,
)

# %% ...........................  workflow step #3 - PCA  ..............................


def PCA_main(avg_dfs, folderinfo, cfg, plot_panel_instance):
    """PCA on features of interest"""

    # unpack
    results_dir = folderinfo["results_dir"]
    PCA_custom_scatter_PCs = cfg["PCA_custom_scatter_PCs"]
    PCA_bins = cfg["PCA_bins"]

    # convert PCA_bins string to list of ints
    if PCA_bins:
        cfg = convert_PCA_bins_to_list(folderinfo, cfg)
        info_string = "Desired PCA bins applied to cycle-bins resulted in:\n|"
        for bin in cfg["PCA_bins"]:
            info_string += str(bin) + "|"
        print(info_string)
    # create the input dataframe
    PCA_df, features = create_PCA_df(avg_dfs, folderinfo, cfg)
    # run the PCA
    PCA_df, PCA_info = run_PCA(PCA_df, features, cfg)
    # save PCA info to xlsx files
    PCA_info_to_xlsx(PCA_df, PCA_info, folderinfo, cfg)
    if PCA_info["number_of_PCs"] < 2:
        PCA_error_message_1 = (
            "\n***********\n! ERROR !\n***********\n"
            + "PC scatters were not generated because there are less than 2 PCs!"
            + "\nThis hints at a low value of PCA_n_components using the "
            + "var-explained approach!"
        )
        print(PCA_error_message_1)
        write_issues_to_textfile(PCA_error_message_1, results_dir)
        return
    # plot barplot of cumulative explained variance
    plot_PCA_barplot(PCA_info, folderinfo, cfg, plot_panel_instance)
    # plot the standard scatterplot
    plot_PCA_scatterplots(
        "Standard", PCA_df, PCA_info, folderinfo, cfg, plot_panel_instance
    )
    # plot custom scatterplots if user wanted
    # => loop over outer separator of PC-duos/trios
    if PCA_custom_scatter_PCs:
        for i in range(
            len(PCA_custom_scatter_PCs.split(PCA_CUSTOM_SCATTER_OUTER_SEPARATOR))
        ):
            plot_PCA_scatterplots(
                "Custom #" + str((i + 1)),
                PCA_df,
                PCA_info,
                folderinfo,
                cfg,
                plot_panel_instance,
            )


def convert_PCA_bins_to_list(folderinfo, cfg):
    """Create PCA bin variable (list of to-be-included bin_num ints) from user input
    (string).
    Approach
    --------
    1. (Note we already removed spaces in group_1_prep). Also initialise a string list that has all bins (, separated) and an empty list
    2. Iterate over string list and extract start and end indices if it's a range (idenfying range by "-")
    3. Concatenate the range to the int list, make sure that you add a 1 to end idx
    to have user input be inclusive
    4. If it's not a range, just add the int to the int list
    5. Now remove all values that are not part of our SC Percentages (use bin_num)
    for this
    """
    # unpack
    results_dir = folderinfo["results_dir"]
    PCA_bins = cfg["PCA_bins"]
    bin_num = cfg["bin_num"]

    PCA_bin_str_list = PCA_bins.split(",")
    PCA_bin_int_list = []
    for bin in PCA_bin_str_list:
        if "-" in bin:
            bin_range = bin.split("-")
            # +1 in indexing of bin_range [1] ensures including user input fully!
            try:
                PCA_bin_int_list = np.concatenate(
                    [
                        PCA_bin_int_list,
                        np.arange(int(bin_range[0]), int(bin_range[1]) + 1),
                    ]
                )
            except ValueError:
                PCA_bins_error_message = (
                    "\n*********\n! ERROR !\n*********\n"
                    + "\nPCA bin range input invalid."
                    + "\nSeems like you used a hyphen without numbers around it."
                    + "\nPlease check your input and try again!"
                )
                print(PCA_bins_error_message)
                write_issues_to_textfile(PCA_bins_error_message, results_dir)
                raise ValueError(PCA_bins_error_message)

        else:
            PCA_bin_int_list = np.concatenate([PCA_bin_int_list, [int(bin)]])
        PCA_bin_int_list = PCA_bin_int_list.astype(int)  # for intersect below
    # Note that we can be sure that this is the same list that was used at the first
    # level since it's the same function used there
    percentages_list = bin_num_to_percentages(bin_num)
    # Now using intersect 1d that returns the sorted unique values that are in both
    # arrays
    cfg["PCA_bins"] = np.intersect1d(PCA_bin_int_list, percentages_list)
    # Just in case... remove everything over 100 (our error above ensures nothing
    # smaller than 0 already)
    cfg["PCA_bins"] = cfg["PCA_bins"][cfg["PCA_bins"] <= 100]
    return cfg


def create_PCA_df(avg_dfs, folderinfo, cfg):
    """Create a ID x ID_COL + features dataframe to be used by PCA"""

    # unpack
    results_dir = folderinfo["results_dir"]
    group_names = folderinfo["group_names"]
    PCA_vars = cfg["PCA_variables"]
    bin_num = cfg["bin_num"]
    PCA_bins = cfg["PCA_bins"]

    PCA_df = pd.DataFrame(data=None)
    # create a list of features for series & dfs (features are vars @ each SC % bin)
    features = []
    for var in PCA_vars:
        if len(PCA_bins) > 0:
            for bin in PCA_bins:
                features.append(var + " " + str(bin))
        else:
            for b in range(bin_num):
                bin_in_percent = int(((1 + b) / bin_num) * 100)
                features.append(var + " " + str(bin_in_percent))

    # for each mouse, create a series to concat to PCA_df
    for g, group_name in enumerate(group_names):
        for ID in pd.unique(avg_dfs[g][ID_COL]):
            this_list = [group_name, ID]
            if len(PCA_bins) > 0:  # empty strings are len = 0
                # first line gives True and False but we need indexes so it's
                # compatible with my (previous) approach of using iloc and lists
                row_idxs = (avg_dfs[g][ID_COL] == ID) & (
                    avg_dfs[g][SC_PERCENTAGE_COL].isin(PCA_bins)
                )
                row_idxs = np.where(row_idxs == True)[0]
            else:
                row_idxs = np.where(avg_dfs[g][ID_COL] == ID)[0]
            for var in PCA_vars:
                joint_col_idx = avg_dfs[g].columns.get_loc(var)
                # this data is of current ID or IDxSC% combo and current feature
                this_data = list(avg_dfs[g].iloc[row_idxs, joint_col_idx].values)
                this_list.extend(this_data)  # so we have ID_COL as 1st value
            this_series = pd.Series(this_list)
            if PCA_df.empty:
                PCA_df = pd.DataFrame(this_series).transpose()
            else:
                # transpose series, transform to df and concat to row-axis
                PCA_df = pd.concat([PCA_df, this_series.to_frame().T], axis=0)
    # this check is me being paranoid but whatever let's just leave it in
    if len(PCA_df.columns) != len(features) + 2:
        critical_PCA_df_error_message = (
            "\n*********\n! ERROR !\n*********\n"
            + "PCA_df creation failed since columns did not match features!\n"
            + "This should never happen, please contact me!"
        )
        print(critical_PCA_df_error_message)
        write_issues_to_textfile(critical_PCA_df_error_message, results_dir)
        raise ValueError(critical_PCA_df_error_message)
    # add colnames after the last mouse (makes concat'ing series 2 df easier)
    PCA_df.columns = [GROUP_COL] + [ID_COL] + features
    return PCA_df, features


def run_PCA(PCA_df, features, cfg):
    """Runs the PCA on a limb's feature (e.g. y or z coordinates)"""

    # unpack
    PCA_n_components = cfg["PCA_n_components"]

    # run
    if 0 < PCA_n_components < 1:  # n_components until % explained var.
        PCA_model = PCA(n_components=PCA_n_components, svd_solver="full")
    else:
        PCA_model = PCA(n_components=int(PCA_n_components))
    x = PCA_df.loc[:, features]
    # standardise here so that EACH FEATURE has mean=0 & std=1
    # => you can check this with np.mean/std
    # => you can also check that this operates on columns (features) and not rows (IDs)
    x = StandardScaler().fit_transform(x)
    # now create a df with the standardised features to add features to it (this gives
    # us the feature_names_in attribute of PCA_model which we will use for PCA info
    # excel files)
    x = pd.DataFrame(x, columns=features)
    # NOTE that this next line changes the PCA_model variable!
    PCs = PCA_model.fit_transform(x)
    # NOTE! we use number_of_PCs from here onwards (in info/plots/etc) to differentiate
    #       between n_components which can be smaller than 1!
    number_of_PCs = np.shape(PCs)[1]
    for i in range(number_of_PCs):
        PCA_df["PC " + str(i + 1)] = PCs[:, i]
    PCA_info = {
        "features": PCA_model.feature_names_in_,
        "number_of_PCs": number_of_PCs,
        "explained_vars": PCA_model.explained_variance_ratio_,
        "eigenvectors": PCA_model.components_,
    }
    return PCA_df, PCA_info


def PCA_info_to_xlsx(PCA_df, PCA_info, folderinfo, cfg):
    """Save three PCA info excel files:
    1) explained_var & eigenvectors of PCs
    2) ID x PC values
    3) top 20 features for each PC using eigenvectors
    """

    # unpack
    results_dir = folderinfo["results_dir"]
    features = PCA_info["features"]  # PCA info!
    number_of_PCs = PCA_info["number_of_PCs"]
    explained_vars = PCA_info["explained_vars"]
    eigenvectors = PCA_info["eigenvectors"]

    # 1) PCs' explained variances & eigenvectors

    #  initialise
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "PCA Info"

    # add column headers
    for pc in range(number_of_PCs):
        sheet[string.ascii_uppercase[pc + 1] + "1"] = "PC " + str(pc + 1)
    # add cell values: explained variance
    sheet.cell(row=2, column=1, value="Explained Var. (%)")
    for pc in range(number_of_PCs):
        sheet.cell(row=2, column=pc + 2, value=round(explained_vars[pc] * 100, 2))
    # add cell values: eigenvectors
    sheet.cell(row=4, column=1, value="Features")
    for i in range(len(features)):
        # row i + 5 because excel starts counting at 1 and: row2=var_exp/row3=empty/
        # row5=feature heading
        sheet.cell(row=i + 5, column=1, value=features[i])
        for pc in range(number_of_PCs):
            # column is pc+2 because we want pc=0 to be in xlsx column 2 etc.
            sheet.cell(row=i + 5, column=pc + 2, value=eigenvectors[pc, i])
    # save
    workbook.save(os.path.join(results_dir, "PCA Info.xlsx"))

    # 2) ID x PC value information
    PCA_ID_info_df = PCA_df[[GROUP_COL, ID_COL] + list(PCA_df.columns[-number_of_PCs:])]
    PCA_ID_info_df.to_excel(os.path.join(results_dir, "PCA ID Info.xlsx"), index=False)

    # 3) Find the biggest 20 values of eigenvectors for each PC
    if len(features) < 20:
        top_number = len(features)
    else:
        top_number = 20
    top_features_df = pd.DataFrame()
    for pc in range(number_of_PCs):
        # argsort returns idxs that would sort in ascending order, so we reverse it
        # (::-1)
        top_features = np.argsort(np.abs(eigenvectors[pc]))[::-1][:top_number]
        top_values = eigenvectors[pc, top_features]
        top_features_df["PC " + str(pc + 1) + " Features"] = features[top_features]
        top_features_df["PC " + str(pc + 1) + " Value"] = top_values
    filename = "PCA Feature Summary.xlsx"
    top_features_df.to_excel(os.path.join(results_dir, filename), index=False)


def plot_PCA_barplot(PCA_info, folderinfo, cfg, plot_panel_instance):
    """Plot a barplot illustrating PCs explained variance"""

    # unpack
    number_of_PCs = PCA_info["number_of_PCs"]  # PCA info!
    explained_vars = PCA_info["explained_vars"]
    results_dir = folderinfo["results_dir"]
    dont_show_plots = cfg["dont_show_plots"]

    # plot
    df = pd.DataFrame(
        data=explained_vars * 100,
        columns=["Explained Variance"],
        index=np.arange(1, PCA_info["number_of_PCs"] + 1),
    )
    cumulative_explained_var = []
    for i in range(number_of_PCs):
        cumulative_explained_var.append(np.sum(explained_vars[: i + 1]) * 100)
    f, ax = plt.subplots(1, 1)
    sns.barplot(df["Explained Variance"], color=PCA_BARPLOT_BARCOLOR, ax=ax)
    ax.set_xlabel("Principal Components")
    ax.set_ylabel("Explained Variance (%)")
    ax.plot(
        cumulative_explained_var,
        color=PCA_BARPLOT_LINECOLOR,
        linestyle=":",
        marker="o",
        label="Cumulative Explained Variance",
    )
    info_string = "PCA Explained Variance"
    save_figures(f, results_dir, info_string)
    # add figure to plot panel figures list
    if dont_show_plots is False:  # -> show plot panel
        plot_panel_instance.figures.append(f)


def plot_PCA_scatterplots(
    which_plot, PCA_df, PCA_info, folderinfo, cfg, plot_panel_instance
):
    """Plot a scatterplot and colour based on group name"""

    # unpack
    group_names = folderinfo["group_names"]
    results_dir = folderinfo["results_dir"]
    group_color_dict = cfg["group_color_dict"]
    PCA_custom_scatter_PCs = cfg["PCA_custom_scatter_PCs"]
    PCA_save_3D_video = cfg["PCA_save_3D_video"]
    dont_show_plots = cfg["dont_show_plots"]
    legend_outside = cfg["legend_outside"]

    # number of PCs & custom flag to control standard/custom plots
    number_of_PCs = PCA_info["number_of_PCs"]  # PCA info!
    custom_flag = False
    if "Custom" in which_plot:
        custom_flag = True
        # first, extract PC numbers from the cfg string and test input validity
        # => important: we need to subtract 1 after extracting from the string because
        #               PC numbers are 1-indexed
        PC_idx = int(which_plot[-1]) - 1  # here!
        all_scatter_PC_sets = PCA_custom_scatter_PCs.split(
            PCA_CUSTOM_SCATTER_OUTER_SEPARATOR
        )
        this_scatter_PC_set = all_scatter_PC_sets[PC_idx]
        scatter_PC_nums = this_scatter_PC_set.split(PCA_CUSTOM_SCATTER_INNER_SEPARATOR)
        scatter_PC_nums = [int(this_PC) for this_PC in scatter_PC_nums]
        # Important Note!
        # => below test must be done on "real" number of PCs of full PCA from PCA info
        if np.max(scatter_PC_nums) > number_of_PCs:
            invalid_custom_PC_scatter_message = (
                "\n*********\n! ERROR !\n*********\n"
                + "\nCustom PCA scatterplot configuration invalid for:\n"
                + PCA_custom_scatter_PCs[PC_idx]
                + "!\nThe PC you wanted was not computed by PCA!"
            )
            print(invalid_custom_PC_scatter_message)
            write_issues_to_textfile(invalid_custom_PC_scatter_message, results_dir)
            return
        # after the test we can overwrite "real" number of PCs of full PCA here since
        # we only use this value to know if we also do a 3D plot or not from here on
        number_of_PCs = len(scatter_PC_nums)

    # prepare PC strings based on which plot we are making and extract col_idxs from
    # PCA_df for scatter
    if custom_flag:
        PC1_str = "PC " + str(scatter_PC_nums[0])
        PC2_str = "PC " + str(scatter_PC_nums[1])
        if number_of_PCs > 2:
            PC3_str = "PC " + str(scatter_PC_nums[2])
    else:
        PC1_str = "PC 1"
        PC2_str = "PC 2"
        if number_of_PCs > 2:
            PC3_str = "PC 3"
    PC1_col_idx = PCA_df.columns.get_loc(PC1_str)
    PC2_col_idx = PCA_df.columns.get_loc(PC2_str)
    if number_of_PCs > 2:
        PC3_col_idx = PCA_df.columns.get_loc(PC3_str)

    # loop over groups, and call scatter 3x so we can have a correct legend
    f, ax = plt.subplots(1, 1)
    if number_of_PCs > 2:
        f_3d = plt.figure()
        ax_3d = f_3d.add_subplot(111, projection="3d")
    for g, group_name in enumerate(group_names):
        row_idxs = np.where(PCA_df[GROUP_COL] == group_name)[0]
        x = PCA_df.iloc[row_idxs, PC1_col_idx].values
        y = PCA_df.iloc[row_idxs, PC2_col_idx].values
        ax.scatter(x, y, color=group_color_dict[group_name], label=group_name)
        if number_of_PCs > 2:
            z = PCA_df.iloc[row_idxs, PC3_col_idx].values
            ax_3d.scatter(
                x, y, z, color=group_color_dict[group_name], s=60, label=group_name
            )
    # legend, labels & info-string for title/save
    if legend_outside is True:
        ax.legend(group_names, loc="center left", bbox_to_anchor=(1, 0.5))
    elif legend_outside is False:
        ax.legend(group_names)
    ax.set_xlabel(PC1_str)
    ax.set_ylabel(PC2_str)
    info_string = which_plot + " 2D PCA Scatterplot"
    ax.set_title(info_string)
    if number_of_PCs > 2:  # 3d scatterplot
        ax_3d.view_init(30, 125)
        # legend adjustments
        if legend_outside is True:
            ax_3d.legend(group_names, loc="center right", bbox_to_anchor=(0, 0.5))
        elif legend_outside is False:
            ax_3d.legend(group_names)
        ax_3d.set_xlabel(PC1_str)
        ax_3d.set_ylabel(PC2_str)
        ax_3d.set_zlabel(PC3_str)
        info_string_3d = which_plot + " 3D PCA Scatterplot"
        ax_3d.set_title(info_string_3d)
    save_figures(f, results_dir, info_string)
    # add figure to plot panel figures list
    if dont_show_plots is False:  # -> show plot panel
        plot_panel_instance.figures.append(f)

    # 3d scatterplot image file
    if number_of_PCs > 2:
        save_figures(f_3d, results_dir, info_string_3d)
        # add figure to plot panel figures list
        if dont_show_plots is False:  # -> show plot panel
            plot_panel_instance.figures.append(f_3d)

        # 3d scatterplot rotating video file
        if PCA_save_3D_video:

            def init():  # create the animation
                return (f_3d,)

            def animate(frame):
                ax_3d.view_init(elev=10, azim=frame)
                return (f_3d,)

            anim = FuncAnimation(
                f_3d, animate, frames=np.arange(0, 360, 1), interval=20, blit=True
            )
            writervideo = FFMpegWriter(fps=30)  # save to m4 using ffmpeg writer
            anim.save(
                os.path.join(results_dir, info_string_3d + ".mp4"), writer=writervideo
            )
