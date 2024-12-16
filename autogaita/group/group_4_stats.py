# %% imports
from autogaita.resources.utils import bin_num_to_percentages
from autogaita.group.group_utils import (
    check_mouse_conversion,
    save_figures,
    ytickconvert_mm_to_cm,
    ylabel_velocity_and_acceleration,
    write_issues_to_textfile,
)
import os
import sys
import pandas as pd
import numpy as np
import string
import openpyxl
from sklearn.utils import shuffle
import matplotlib.pyplot as plt
from scipy import stats
import pingouin as pg

# %% constants
from autogaita.resources.constants import INFO_TEXT_WIDTH, ID_COL, SC_PERCENTAGE_COL

from autogaita.group.group_constants import (
    STATS_TXT_FILENAME,
    GROUP_COL,
    CONTRASTS_COL,
    CONTRAST_SPLIT_STR,
    TTEST_MASK_THRESHOLD,  # STATS
    TTEST_P_COL,
    TTEST_T_COL,
    TTEST_MASK_COL,
    CLUSTER_TMASS_COL,
    CLUSTER_P_COL,
    CLUSTER_MASK_COL,
    MULTCOMP_EXCEL_FILENAME_1,
    MULTCOMP_EXCEL_FILENAME_2,
    MULTCOMP_RESULT_TYPES,
    MULTCOMP_RESULT_P_IDENTIFIER,
    MULTCOMP_RESULT_SPLIT_STR,
    MULTCOMP_EXCEL_COLS,
    STATS_PLOT_LEGEND_SIZE,  # PLOTS
    STATS_PLOTS_SUPLABEL_SIZE,
    BOX_COLOR,
    BOX_ALPHA,
    STD_ALPHA,
    STD_LW,
)

# %% ....................  workflow step #4 - statistics  ..............................


# %% ................  local functions #1 - prepare statistics  ........................


def create_stats_df(avg_dfs, folderinfo, cfg):
    """Create a df in format used by (both) our stats approaches.
    Shape: ID*SC% x features (i.e. all IDs' averages concatenated along rows)
    """

    # unpack
    group_names = folderinfo["group_names"]
    bin_num = cfg["bin_num"]

    for g, group_name in enumerate(group_names):
        avg_dfs[g][GROUP_COL] = group_name
        if g == 0:
            stats_df = avg_dfs[g]
        else:
            stats_df = pd.concat([stats_df, avg_dfs[g]], axis=0)
    # note that we have to repeat percentages_list to match the number of rows in
    # stats_df
    percentages_list = bin_num_to_percentages(bin_num)
    how_often_to_repeat_percentages_list = len(stats_df) // len(percentages_list)
    stats_df[SC_PERCENTAGE_COL] = np.tile(
        percentages_list, how_often_to_repeat_percentages_list
    )
    return stats_df


# %% ...........  local functions #2 - cluster-extent permutation test  ................


# ...............................  main function  ......................................
def cluster_extent_test(
    stats_df, g_avg_dfs, g_std_dfs, stats_var, folderinfo, cfg, plot_panel_instance
):
    """Main function running a cluster-extent permutation test of N contrasts for a
    given dependent variable
    """

    # unpack
    permutation_number = cfg["permutation_number"]
    stats_threshold = cfg["stats_threshold"]
    # initialise the text file
    initial_stats_textfile(stats_var, "Permutation Test", folderinfo)
    # true observed
    trueobs_df = stats_df.copy()
    trueobs_results_df = initialise_results_df(folderinfo, cfg)
    trueobs_results_df = compute_first_level_results(
        trueobs_df, trueobs_results_df, stats_var, folderinfo
    )
    # permutation
    max_tmass = np.zeros(permutation_number)  # tmass
    for p in range(permutation_number):
        permuted_df = permute_true_observed_df(trueobs_df, cfg)
        permuted_results_df = initialise_results_df(folderinfo, cfg)
        permuted_results_df = compute_first_level_results(
            permuted_df, permuted_results_df, stats_var, folderinfo
        )
        # max tmass
        max_tmass[p] = max(permuted_results_df[CLUSTER_TMASS_COL])
        sys.stdout.write(
            "\r*************** Permuting "
            + stats_var
            + ": "
            + str(p + 1)
            + "/"
            + str(permutation_number)
            + " ***************"
        )
        sys.stdout.flush()
    # assign final p values of true observed cluster sizes
    trueobs_results_df = test_trueobs_clusters(
        trueobs_results_df, max_tmass, permutation_number, stats_threshold
    )
    # print & save exact numerical results (significant SC % clusters) to a textfile
    save_stats_summary_to_text(
        trueobs_results_df,
        "Permutation Test",
        folderinfo,
        cfg,
    )
    # plot results
    plot_permutation_test_results(
        g_avg_dfs,
        g_std_dfs,
        trueobs_results_df,
        stats_var,
        folderinfo,
        cfg,
        plot_panel_instance,
    )


# ................................  preparation  .......................................
def initialise_results_df(folderinfo, cfg):
    """Initialise a results df for permutation test."""

    # unpack
    contrasts = folderinfo["contrasts"]
    bin_num = cfg["bin_num"]

    initial_contrasts_list = []
    for contrast in contrasts:
        initial_contrasts_list.extend([contrast] * bin_num)
    results_df = pd.DataFrame(data=initial_contrasts_list, columns=[CONTRASTS_COL])
    results_df[SC_PERCENTAGE_COL] = None
    results_df[TTEST_P_COL] = float(1)
    results_df[TTEST_T_COL] = float(0)
    results_df[TTEST_MASK_COL] = False
    results_df[CLUSTER_TMASS_COL] = 0.0
    return results_df


# ...............................  first-level  ........................................
def compute_first_level_results(stats_df, results_df, stats_var, folderinfo):
    """Compute the results of our first level, i.e., mass ttests"""

    # unpack
    contrasts = folderinfo["contrasts"]

    # populate true observed results df
    idx = 0
    for contrast in contrasts:
        group1 = contrast.split(CONTRAST_SPLIT_STR)[0]
        group2 = contrast.split(CONTRAST_SPLIT_STR)[1]
        # SC percentage & ttest results
        for s, sc_percentage in enumerate(np.unique(stats_df[SC_PERCENTAGE_COL])):
            results_df = run_and_assign_ttest(
                stats_df,
                results_df,
                stats_var,
                contrast,
                group1,
                group2,
                sc_percentage,
                idx,
            )
            # update idx for assigning sc_perc in next iter correctly (function input)
            idx += 1
        # cluster size
        results_df = compute_and_assign_clustersize(results_df, contrast)
    return results_df


def run_and_assign_ttest(
    stats_df, results_df, stats_var, contrast, group1, group2, sc_percentage, idx
):
    """Run ttest for a given pair of groups & a given percentage of the step cycle."""
    # get location of current SC Percentage
    sc_percentage_col_idx = results_df.columns.get_loc(SC_PERCENTAGE_COL)
    results_df.iloc[idx, sc_percentage_col_idx] = sc_percentage
    # extract the two arrays to be tested, test & get its results
    arr1 = extract_variable_array(stats_df, stats_var, group1, sc_percentage)
    arr2 = extract_variable_array(stats_df, stats_var, group2, sc_percentage)
    this_t, this_p = stats.ttest_ind(arr1, arr2)
    # assign the t & p-value and assign significance mask to mask-column
    # ==> this_result_rowidx_mask is all False & one True, with the True being current
    #     contrast & SC Percentage
    this_result_rowidx_mask = (results_df[CONTRASTS_COL] == contrast) & (
        results_df[SC_PERCENTAGE_COL] == sc_percentage
    )
    results_df.loc[this_result_rowidx_mask, TTEST_T_COL] = this_t
    results_df.loc[this_result_rowidx_mask, TTEST_P_COL] = this_p
    if this_p < TTEST_MASK_THRESHOLD:
        results_df.loc[this_result_rowidx_mask, TTEST_MASK_COL] = True
    return results_df


def compute_and_assign_clustersize(results_df, contrast):
    """Compute size of all clusters of a given contrast and assign to results_df."""
    # prepare some variables
    ttest_mask_col_idx = results_df.columns.get_loc(TTEST_MASK_COL)
    ttest_tval_col_idx = results_df.columns.get_loc(TTEST_T_COL)
    tmass_col_idx = results_df.columns.get_loc(CLUSTER_TMASS_COL)
    this_tmass = 0.0  # tmass
    this_cluster_size = 0  # cluster size only used for checking if all ps were sig
    this_cluster_indices = []
    # loop over current contrast, update tmass, cluster size & indices if p was
    # significant
    for i in np.where(results_df[CONTRASTS_COL] == contrast)[0]:
        if results_df.iloc[i, ttest_mask_col_idx] == True:
            this_tmass += abs(results_df.iloc[i, ttest_tval_col_idx])  # tval
            this_cluster_size += 1  # cluster size
            this_cluster_indices.append(i)
            # handle the case of results_df ending with a sig. cluster
            if i == max(np.where(results_df[CONTRASTS_COL] == contrast)[0]):
                results_df.iloc[this_cluster_indices, tmass_col_idx] = this_tmass
        else:
            # if p was not significant, assign the previous cluster & reset our vars
            # => note this else also occurs when we keep having nonsig ts but for those
            #    nothing happens... coding it like this might make it a bit slower than
            #    a more sophisticated conditional logic here but the difference should
            #    be minimal so I just keep it as is
            results_df.iloc[this_cluster_indices, tmass_col_idx] = this_tmass
            this_tmass = 0.0
            this_cluster_size = 0
            this_cluster_indices = []
    # handle case of all ps being significant
    if this_cluster_size == len(np.where(results_df[CONTRASTS_COL] == contrast)[0]):
        results_df.iloc[this_cluster_indices, tmass_col_idx] = this_tmass
    return results_df


def extract_variable_array(df, stats_var, group_name, sc_percentage):
    """Extract an array of the variable we want to test with given ttest."""
    mask = (df[GROUP_COL] == group_name) & (df[SC_PERCENTAGE_COL] == sc_percentage)
    result = df.loc[mask, stats_var]
    return np.asarray(result)


# .........................  shuffle (permute) the true observed  ......................
def permute_true_observed_df(trueobs_df, cfg):
    """Shuffle groups of true observed and return permuted_df which is identical to
    trueobs_df except of GROUP_COL
    """
    # unpack
    bin_num = cfg["bin_num"]

    real_group_labels = list(trueobs_df.loc[0, GROUP_COL])
    permuted_group_labels = shuffle(real_group_labels)
    permuted_group_labels = np.repeat(permuted_group_labels, bin_num)
    permuted_df = trueobs_df.copy(deep=True)  # make sure to use copy here (not df=df)!!
    permuted_df.loc[:, GROUP_COL] = permuted_group_labels
    return permuted_df


# ................................  second-level test  .................................
def test_trueobs_clusters(
    trueobs_results_df, max_tmass, permutation_number, stats_threshold
):
    """Test the true observed cluster sizes against max cluster sizes under null."""
    # prepare stuff
    trueobs_results_df[CLUSTER_P_COL] = None
    trueobs_results_df[CLUSTER_MASK_COL] = False
    cluster_p_col_idx = trueobs_results_df.columns.get_loc(CLUSTER_P_COL)
    cluster_tmass_col_idx = trueobs_results_df.columns.get_loc(CLUSTER_TMASS_COL)
    cluster_mask_col_idx = trueobs_results_df.columns.get_loc(CLUSTER_MASK_COL)
    # loop over results, check each clustersize, assign final p value & mask
    for i in range(len(trueobs_results_df)):
        # tmass
        this_p = (
            sum(max_tmass >= trueobs_results_df.iloc[i, cluster_tmass_col_idx])
            / permutation_number
        )
        trueobs_results_df.iloc[i, cluster_p_col_idx] = this_p
        if this_p < stats_threshold:
            trueobs_results_df.iloc[i, cluster_mask_col_idx] = True
    return trueobs_results_df


# ...................................  plot results  ...................................
def plot_permutation_test_results(
    g_avg_dfs,
    g_std_dfs,
    trueobs_results_df,
    stats_var,
    folderinfo,
    cfg,
    plot_panel_instance,
):
    """Plot a Nx1 or N/2x2 figure of our contrasts' permutation test results."""

    # unpack
    group_names = folderinfo["group_names"]
    results_dir = folderinfo["results_dir"]
    contrasts = folderinfo["contrasts"]
    sampling_rate = cfg["sampling_rate"]
    bin_num = cfg["bin_num"]
    group_color_dict = cfg["group_color_dict"]
    plot_SE = cfg["plot_SE"]
    tracking_software = cfg["tracking_software"]
    feature = stats_var.split(" ")[-1]
    dont_show_plots = cfg["dont_show_plots"]
    legend_outside = cfg["legend_outside"]

    if len(contrasts) > 3:  # if we have 4 groups or more, N/2x2 subplot layout
        f, ax = plt.subplots(int(round(len(contrasts) / 2)), 2, layout="constrained")
        ax = ax.ravel()
    else:
        f, ax = plt.subplots(len(contrasts), 1, layout="constrained")
    x = np.linspace(0, 100, bin_num)
    for c, contrast in enumerate(contrasts):
        # prepare group strings and (importantly!) index of current groups from _NAMES
        groups = [group_name for group_name in contrast.split(CONTRAST_SPLIT_STR)]
        group_indices = [group_names.index(group_name) for group_name in groups]
        # plot observed g_avgs & g_stds
        for g, group_name in enumerate(groups):
            # group_idx is important. it correctly indexes dfs (!!) as well as colour!
            group_idx = group_indices[g]
            y_col = g_avg_dfs[group_idx].columns.get_loc(stats_var)
            y = g_avg_dfs[group_idx].iloc[:, y_col]
            if plot_SE:
                std = g_std_dfs[group_idx].iloc[:, y_col] / np.sqrt(
                    g_std_dfs[group_idx]["N"][0]
                )
            else:
                std = g_std_dfs[group_idx].iloc[:, y_col]
            this_color = group_color_dict[group_name]
            if type(ax) == np.ndarray:  # so we can do 2-group contrasts
                ax[c].plot(x, y, color=this_color, label=group_name, zorder=1)
                ax[c].fill_between(
                    x,
                    y - std,
                    y + std,
                    color=this_color,
                    alpha=STD_ALPHA,
                    lw=STD_LW,
                    zorder=1,
                )
            else:
                ax.plot(x, y, color=this_color, label=group_name, zorder=1)
                ax.fill_between(
                    x,
                    y - std,
                    y + std,
                    color=this_color,
                    alpha=STD_ALPHA,
                    lw=STD_LW,
                    zorder=1,
                )
        # adjust legend & convert to cm (if needed) before plotting clusters
        if type(ax) == np.ndarray:
            if legend_outside is True:
                ax[c].legend(
                    fontsize=STATS_PLOT_LEGEND_SIZE,
                    loc="center left",
                    bbox_to_anchor=(1, 0.5),
                )
            elif legend_outside is False:
                ax[c].legend(fontsize=STATS_PLOT_LEGEND_SIZE)
            if check_mouse_conversion(feature, cfg, stats_var=stats_var):
                ytickconvert_mm_to_cm(ax[c])
        else:
            if legend_outside is True:
                ax.legend(
                    fontsize=STATS_PLOT_LEGEND_SIZE + 4,
                    loc="center left",
                    bbox_to_anchor=(1, 0.5),
                )
            elif legend_outside is False:
                ax.legend(fontsize=STATS_PLOT_LEGEND_SIZE + 4)
            if check_mouse_conversion(feature, cfg, stats_var=stats_var):
                ytickconvert_mm_to_cm(ax)
        # plot significant clusters
        # => note that clusters is a list of list with idxs between 0 & bin_num-1
        clusters = extract_all_clusters(trueobs_results_df, contrast)
        if type(ax) == np.ndarray:
            ymin = ax[c].get_ylim()[0]
            ymax = ax[c].get_ylim()[1]
        else:
            ymin = ax.get_ylim()[0]
            ymax = ax.get_ylim()[1]
        for cluster in x[clusters]:  # index x with clusters == cluster has correct val
            x_coords = [cluster[0], cluster[1], cluster[1], cluster[0]]
            y_coords = [ymin, ymin, ymax, ymax]
            if type(ax) == np.ndarray:
                ax[c].fill(
                    x_coords,
                    y_coords,
                    color=BOX_COLOR,
                    alpha=BOX_ALPHA,
                    lw=STD_LW,
                    zorder=0,
                )
            else:
                ax.fill(
                    x_coords,
                    y_coords,
                    color=BOX_COLOR,
                    alpha=BOX_ALPHA,
                    lw=STD_LW,
                    zorder=0,
                )
    f.supxlabel("Percentage", fontsize=STATS_PLOTS_SUPLABEL_SIZE)
    # ylabels depend on whether we converted mm to cm and on the feature
    # code below calls the ylabel function for all possible cases:
    # 1) (DLC only) converted velocity & acceleration
    # 2) (DLC only) converted x/y coordinates
    # 3) (non-converted) angular velocity & acceleration
    # 4) (non-converted) x(DLC)/Y(Universal 3D) velocity & acceleration
    # 5) (non-converted) x(DLC)/Y(Universal 3D) coordinates
    if check_mouse_conversion(feature, cfg, stats_var=stats_var):
        if feature in ["Velocity", "Acceleration"]:
            f.supylabel(
                ylabel_velocity_and_acceleration(feature, "x in cm", sampling_rate),
                fontsize=STATS_PLOTS_SUPLABEL_SIZE,
            )
        else:
            f.supylabel(feature + " (cm)", fontsize=STATS_PLOTS_SUPLABEL_SIZE)
    else:
        if feature in ["Velocity", "Acceleration"]:
            if "Angle" in stats_var:
                unit = "degrees"
            else:
                if tracking_software in ["DLC", "SLEAP"]:
                    unit = "x in pixels"
                elif tracking_software == "Universal 3D":
                    unit = "Y in (your units)"
            f.supylabel(
                ylabel_velocity_and_acceleration(feature, unit, sampling_rate),
                fontsize=STATS_PLOTS_SUPLABEL_SIZE,
            )
        else:
            f.supylabel(feature, fontsize=STATS_PLOTS_SUPLABEL_SIZE)
    figure_file_string = stats_var + " - Cluster-extent Test"
    f.suptitle(figure_file_string, fontsize=STATS_PLOTS_SUPLABEL_SIZE)
    save_figures(f, results_dir, figure_file_string)

    # add figure to plot panel figures list
    if dont_show_plots is False:  # -> show plot panel
        plot_panel_instance.figures.append(f)


def extract_all_clusters(trueobs_results_df, contrast):
    """Find indices of all (perm.) significant clusters"""
    contrast_mask = trueobs_results_df[CONTRASTS_COL] == contrast
    significance_mask = list(trueobs_results_df.loc[contrast_mask, CLUSTER_MASK_COL])
    all_clusters = []
    cluster = []
    for i, mask in enumerate(significance_mask):
        if mask == True:
            if len(cluster) == 0:
                cluster.append(i)
            if i == (len(significance_mask) - 1):
                cluster.append(i)
            else:
                if significance_mask[i + 1] == False:
                    cluster.append(i)
        if len(cluster) == 2:
            all_clusters.append(cluster)
            cluster = []
    return all_clusters


# %% ...................  local functions #3 - RM/Mixed-ANOVA  .........................


# ...............................  sanity check . ......................................
def anova_design_sanity_check(stats_df, folderinfo, cfg):
    """Sanity check the anova_design input of the user based on stats_df's IDs"""

    # unpack
    anova_design = cfg["anova_design"]
    results_dir = folderinfo["results_dir"]

    # get IDs for each group to check if they are unique
    group_IDs = stats_df.groupby("Group")["ID"].unique()
    ID_list = []
    for this_groups_IDs in group_IDs:
        for ID in this_groups_IDs:
            ID_list.append(ID)
    ID_list = [str(IDs) for IDs in ID_list]
    unique_ID_list = list(set(ID_list))

    # Mixed ANOVA - no duplicate IDs across groups!
    if anova_design == "Mixed ANOVA":
        if len(ID_list) == len(unique_ID_list):  # check passed
            return True
        else:
            mixed_anova_error_message = (
                "\n*********\n! ERROR !\n*********\n"
                + "\nANOVA design seems wrong - skipping ANOVA!"
                + "\nMixed ANOVA requires unique IDs across groups & we found "
                + "duplicates!"
                + "\n\nIDs were:\n"
                + str(group_IDs)
            )
            print(mixed_anova_error_message)
            write_issues_to_textfile(mixed_anova_error_message, results_dir)
            return False
    # RM ANOVA - IDs in each group must be the same!
    elif anova_design == "RM ANOVA":
        if len(ID_list) != len(unique_ID_list):  # check passed
            # Bonus - inform user about which IDs were valid (have data)
            # ==> based on pingouin's approach of removing all IDs that do not have
            #     data in all conditions (see https://pingouin-stats.org/build/
            #     html/generated/pingouin.rm_anova.html under "Missing values...")
            valid_IDs = []
            group_number = len(group_IDs.index)
            for ID in unique_ID_list:
                ID_count = ID_list.count(ID)
                if ID_count == group_number:
                    valid_IDs.append(ID)
            rm_anova_info_message = (
                "\n********\n! INFO !\n********\n"
                + "\nFollowing IDs with valid data in all conditions after first-level "
                + "analyses will be included in RM ANOVA:\n\n"
                + str(valid_IDs)
            )
            print(rm_anova_info_message)
            write_issues_to_textfile(rm_anova_info_message, results_dir)
            return True
        else:
            rm_anova_error_message = (
                "\n*********\n! ERROR !\n*********\n"
                + "\nANOVA design seems wrong - skipping ANOVA!"
                + "\nRM ANOVA requires IDs to be present in all groups & we found "
                + "only unique IDs!"
                + "\n\nIDs were:\n"
                + str(group_IDs)
            )
            print(rm_anova_error_message)
            write_issues_to_textfile(rm_anova_error_message, results_dir)
            return False


# ...............................  main function  ......................................
def ANOVA_main(
    stats_df, g_avg_dfs, g_std_dfs, stats_var, folderinfo, cfg, plot_panel_instance
):
    """Perform a two-way RM-ANOVA with the factors group (between or within) & SC
    percentage (within) on a given dependent variable
    """

    # unpack
    anova_design = cfg["anova_design"]

    # initialise text file
    initial_stats_textfile(stats_var, anova_design, folderinfo)

    # run the 1-way RM or Mixed ANOVA
    # => note pingouin checks both or sphericity by default
    ANOVA_result = run_ANOVA(stats_df, stats_var, cfg)

    # run Tukeys for pairwise comparisons
    # => always running multiple comparison tests as well. see Prism's doc for why
    # => https://www.graphpad.com/guides/prism/latest/statistics/
    #    stat_relationship_between_overall_a.htm
    multcomp_df = multcompare_SC_Percentages(stats_df, stats_var, folderinfo, cfg)

    # save results to text, excel and image files
    save_stats_summary_to_text(multcomp_df, anova_design, folderinfo, cfg, ANOVA_result)
    save_multcomp_pvalues_to_excel(multcomp_df, stats_var, folderinfo, cfg)
    plot_multcomp_results(
        g_avg_dfs,
        g_std_dfs,
        multcomp_df,
        stats_var,
        folderinfo,
        cfg,
        plot_panel_instance,
    )


# ..............................  main ANOVA  computation  .............................
def run_ANOVA(stats_df, stats_var, cfg):
    """Run the RM-ANOVA using pingouin"""

    # unpack
    anova_design = cfg["anova_design"]
    if "2-way" in anova_design:  # not really 2-way, see below
        if "RM" in anova_design:
            factor_1_col = cfg["factor_1_col"]
            factor_2_col = cfg["factor_2_col"]
        elif "Mixed" in anova_design:
            within_factor_col = cfg["within_factor_col"]
            between_factor_col = cfg["between_factor_col"]

    # ..............  NOTE FOR MYSELF @ 13.12.2024 before releasing v1  ................
    # => 2-way ANOVAs are 3-way ANOVAs with 2 within and 1 between subjects factor
    #    really
    # => Not supported by Pinguin, likely need to switch to statsmodel's linear mixed
    #    effects models if I want to support it
    # => There is a branch with the it being almost developed - I stopped there with
    #    tests of valid user input (@ line 677 of definefeatures_window) which did not
    #    behave as expected
    # => Still decided to leave these lines here as they are because they do not
    #    affect anything really
    if anova_design == "RM ANOVA":
        result = stats_df.rm_anova(
            dv=stats_var, within=[SC_PERCENTAGE_COL, GROUP_COL], subject=ID_COL
        )
    elif anova_design == "Mixed ANOVA":
        result = stats_df.mixed_anova(
            dv=stats_var,
            within=SC_PERCENTAGE_COL,
            between=GROUP_COL,
            subject=ID_COL,
        )
    elif anova_design == "2-way RM ANOVA":
        result = stats_df.rm_anova(
            dv=stats_var, within=[factor_1_col, factor_2_col], subject=ID_COL
        )
    elif anova_design == "2-way Mixed ANOVA":
        result = stats_df.mixed_anova(
            dv=stats_var,
            within=within_factor_col,
            between=between_factor_col,
            subject=ID_COL,
        )
    return result


# .............................  multiple comparison test  .............................
def multcompare_SC_Percentages(stats_df, stats_var, folderinfo, cfg):
    """Perform multiple comparison test if the ANOVA's interaction was significant.
    Do a separate multcomp test for each SC % bin."""

    # unpack
    group_names = folderinfo["group_names"]
    contrasts = folderinfo["contrasts"]
    bin_num = cfg["bin_num"]
    stats_threshold = cfg["stats_threshold"]

    # prepare multcomp dataframe where we'll store results
    multcomp_df = pd.DataFrame(
        data=np.unique(stats_df[SC_PERCENTAGE_COL]),
        index=range(bin_num),
        columns=[SC_PERCENTAGE_COL],
    )
    result_cols = []  # initialise cols based on different result types and contrasts
    for result_type in MULTCOMP_RESULT_TYPES:
        for contrast in contrasts:
            result_cols.append(result_type + MULTCOMP_RESULT_SPLIT_STR + contrast)
    multcomp_df = pd.concat(
        [
            multcomp_df,
            pd.DataFrame(data=None, index=range(bin_num), columns=result_cols),
        ],
        axis=1,
    )
    # loop over SC Percentages & first prepare depvar_values of current SC % for testing
    # => see below for another 3 for loops when building multcomp_df
    for sc_perc in np.unique(stats_df[SC_PERCENTAGE_COL]):  # loop 1: sc percentage
        depvar_values = [[] for _ in range(len(group_names))]
        for g, group_name in enumerate(group_names):
            sc_perc_condition = stats_df[SC_PERCENTAGE_COL] == sc_perc
            group_condition = stats_df[GROUP_COL] == group_name
            mask = sc_perc_condition & group_condition
            depvar_values[g] = stats_df.loc[mask, stats_var].to_numpy()
        # perform the multcomps test and extract p values
        # =>
        result = stats.tukey_hsd(*depvar_values)  # using * for group_num flexibility
        test_stat = result.statistic
        ps = result.pvalue
        CI = result.confidence_interval(1 - stats_threshold)
        CI_low = CI[0]
        CI_high = CI[1]
        # assign Tukey results to multcomp results df
        # ==> see TukeyHSDResult class of scipy, according to their doc:
        #     "The element at index (i, j) is the p-value for the comparison between
        #     groups i and j." - so i & j matches contrasts as well as ps!
        # ==> This is the case for test_stat and CI_low/high as well
        sc_perc_row_idx = np.where(multcomp_df[SC_PERCENTAGE_COL] == sc_perc)[0][0]
        for i in range(len(group_names)):  # loop 2: group 1
            for j in range(i + 1, len(group_names)):  # loop 3: group 2
                for result_type in MULTCOMP_RESULT_TYPES:  # loop 4: result type
                    # note that cols were initialised appropriately above
                    result_col_idx = multcomp_df.columns.get_loc(
                        result_type
                        + MULTCOMP_RESULT_SPLIT_STR
                        + group_names[i]
                        + CONTRAST_SPLIT_STR
                        + group_names[j]
                    )
                    # strings here have to match list in group_constants
                    if result_type == "q":
                        multcomp_df.iloc[sc_perc_row_idx, result_col_idx] = test_stat[
                            i, j
                        ]
                    # global constant var for p val identifier since used more than once
                    elif result_type == MULTCOMP_RESULT_P_IDENTIFIER:
                        multcomp_df.iloc[sc_perc_row_idx, result_col_idx] = ps[i, j]
                    elif result_type == "CI low":
                        multcomp_df.iloc[sc_perc_row_idx, result_col_idx] = CI_low[i, j]
                    elif result_type == "CI high":
                        multcomp_df.iloc[sc_perc_row_idx, result_col_idx] = CI_high[
                            i, j
                        ]
    return multcomp_df


# ...................................  plot results  ...................................
def plot_multcomp_results(
    g_avg_dfs, g_std_dfs, multcomp_df, stats_var, folderinfo, cfg, plot_panel_instance
):
    """Plot an Nx1 figure of N contrasts' multiple comparison results."""

    # unpack
    group_names = folderinfo["group_names"]
    results_dir = folderinfo["results_dir"]
    contrasts = folderinfo["contrasts"]
    sampling_rate = cfg["sampling_rate"]
    bin_num = cfg["bin_num"]
    group_color_dict = cfg["group_color_dict"]
    stats_threshold = cfg["stats_threshold"]
    plot_SE = cfg["plot_SE"]
    tracking_software = cfg["tracking_software"]
    feature = stats_var.split(" ")[-1]
    dont_show_plots = cfg["dont_show_plots"]
    legend_outside = cfg["legend_outside"]

    f, ax = plt.subplots(len(contrasts), 1, layout="constrained")
    x = np.linspace(0, 100, bin_num)
    for c, contrast in enumerate(contrasts):
        # prepare group strings and (importantly!) index of current groups from _NAMES
        groups = [group_name for group_name in contrast.split(CONTRAST_SPLIT_STR)]
        group_indices = [group_names.index(group_name) for group_name in groups]
        # plot observed g_avgs & g_stds
        for g, group_name in enumerate(groups):
            # group_idx is important. it correctly indexes dfs (!!) as well as colour!
            group_idx = group_indices[g]
            y_col = g_avg_dfs[group_idx].columns.get_loc(stats_var)
            y = g_avg_dfs[group_idx].iloc[:, y_col]
            if plot_SE:
                std = g_std_dfs[group_idx].iloc[:, y_col] / np.sqrt(
                    g_std_dfs[group_idx]["N"][0]
                )
            else:
                std = g_std_dfs[group_idx].iloc[:, y_col]
            this_color = group_color_dict[group_name]
            if type(ax) == np.ndarray:  # so we can do a 2-way contrast
                ax[c].plot(x, y, color=this_color, label=group_name, zorder=1)
                ax[c].fill_between(
                    x,
                    y - std,
                    y + std,
                    color=this_color,
                    alpha=STD_ALPHA,
                    lw=STD_LW,
                    zorder=1,
                )
            else:
                ax.plot(x, y, color=this_color, label=group_name, zorder=1)
                ax.fill_between(
                    x,
                    y - std,
                    y + std,
                    color=this_color,
                    alpha=STD_ALPHA,
                    lw=STD_LW,
                    zorder=1,
                )
        # adjust legend & convert to cm (if needed) before plotting clusters
        if type(ax) == np.ndarray:
            if legend_outside is True:
                ax[c].legend(
                    fontsize=STATS_PLOT_LEGEND_SIZE,
                    loc="center left",
                    bbox_to_anchor=(1, 0.5),
                )
            elif legend_outside is False:
                ax[c].legend(fontsize=STATS_PLOT_LEGEND_SIZE)
            if check_mouse_conversion(feature, cfg, stats_var=stats_var):
                ytickconvert_mm_to_cm(ax[c])
        else:
            if legend_outside is True:
                ax.legend(
                    fontsize=STATS_PLOT_LEGEND_SIZE + 4,
                    loc="center left",
                    bbox_to_anchor=(1, 0.5),
                )
            elif legend_outside is False:
                ax.legend(fontsize=STATS_PLOT_LEGEND_SIZE + 4)
            if check_mouse_conversion(feature, cfg, stats_var=stats_var):
                ytickconvert_mm_to_cm(ax)
        # plot significant clusters
        clusters = extract_multcomp_significance_clusters(
            multcomp_df, contrast, stats_threshold
        )
        if type(ax) == np.ndarray:
            ymin = ax[c].get_ylim()[0]
            ymax = ax[c].get_ylim()[1]
        else:
            ymin = ax.get_ylim()[0]
            ymax = ax.get_ylim()[1]
        for cluster in x[clusters]:  # index x with clusters == cluster has correct val
            x_coords = [cluster[0], cluster[1], cluster[1], cluster[0]]
            y_coords = [ymin, ymin, ymax, ymax]
            if type(ax) == np.ndarray:
                ax[c].fill(
                    x_coords,
                    y_coords,
                    color=BOX_COLOR,
                    alpha=BOX_ALPHA,
                    lw=STD_LW,
                    zorder=0,
                )
            else:
                ax.fill(
                    x_coords,
                    y_coords,
                    color=BOX_COLOR,
                    alpha=BOX_ALPHA,
                    lw=STD_LW,
                    zorder=0,
                )
        if type(ax) == np.ndarray:
            # legend adjustments
            if legend_outside is True:
                ax[c].legend(
                    fontsize=STATS_PLOT_LEGEND_SIZE,
                    loc="center left",
                    bbox_to_anchor=(1, 0.5),
                )
            elif legend_outside is False:
                ax[c].legend(fontsize=STATS_PLOT_LEGEND_SIZE)
        else:
            # legend adjustments
            if legend_outside is True:
                ax.legend(
                    fontsize=STATS_PLOT_LEGEND_SIZE + 4,
                    loc="center left",
                    bbox_to_anchor=(1, 0.5),
                )
            elif legend_outside is False:
                ax.legend(fontsize=STATS_PLOT_LEGEND_SIZE + 4)
    f.supxlabel("Percentage", fontsize=STATS_PLOTS_SUPLABEL_SIZE)
    # ylabels depend on whether we converted mm to cm and on the feature
    # code below calls the ylabel function for all possible cases:
    # 1) (DLC only) converted velocity & acceleration
    # 2) (DLC only) converted x/y coordinates
    # 3) (non-converted) angular velocity & acceleration
    # 4) (non-converted) x(DLC)/Y(Universal 3D) velocity & acceleration
    # 5) (non-converted) x(DLC)/Y(Universal 3D) coordinates
    if check_mouse_conversion(feature, cfg, stats_var=stats_var):
        if feature in ["Velocity", "Acceleration"]:
            f.supylabel(
                ylabel_velocity_and_acceleration(feature, "x in cm", sampling_rate),
                fontsize=STATS_PLOTS_SUPLABEL_SIZE,
            )
        else:
            f.supylabel(feature + " (cm)", fontsize=STATS_PLOTS_SUPLABEL_SIZE)
    else:
        if feature in ["Velocity", "Acceleration"]:
            if "Angle" in stats_var:
                unit = "degrees"
            else:
                if tracking_software in ["DLC", "SLEAP"]:
                    unit = "x in pixels"
                elif tracking_software == "Universal 3D":
                    unit = "Y in (your units)"
            f.supylabel(
                ylabel_velocity_and_acceleration(feature, unit, sampling_rate),
                fontsize=STATS_PLOTS_SUPLABEL_SIZE,
            )
        else:
            f.supylabel(feature, fontsize=STATS_PLOTS_SUPLABEL_SIZE)
    figure_file_string = stats_var + " - Tukey's Multiple Comparison Test"
    f.suptitle(
        figure_file_string,
        fontsize=STATS_PLOTS_SUPLABEL_SIZE,
    )
    save_figures(f, results_dir, figure_file_string)

    # add figure to plot panel figures list
    if dont_show_plots is False:  # -> show plot panel
        plot_panel_instance.figures.append(f)


def extract_multcomp_significance_clusters(multcomp_df, contrast, stats_threshold):
    """Extract clusters of significance after multiple comparison test"""
    # the df structure of this is different to permutation results df so we have to do
    # something slightly different here too
    significance_mask = (
        multcomp_df[MULTCOMP_RESULT_P_IDENTIFIER + MULTCOMP_RESULT_SPLIT_STR + contrast]
        < stats_threshold
    )
    all_clusters = []
    cluster = []
    for i, mask in enumerate(significance_mask):
        if mask == True:
            if len(cluster) == 0:
                cluster.append(i)
            if i == (len(significance_mask) - 1):
                cluster.append(i)
            else:
                if significance_mask[i + 1] == False:
                    cluster.append(i)
        # NU - to handle only 1 sig p value:
        # if len(cluster) == 1 or len(cluster) == 2:....
        # and then handle in plot multcomp with a dashed vertical line or so instead of # a transparent box for
        if len(cluster) == 2:
            all_clusters.append(cluster)
            cluster = []
    return all_clusters


# ...........................  utils - save to txt & excel  ............................
def initial_stats_textfile(stats_var, which_test, folderinfo):
    """Initialise the stats results based on current contrast & analysis"""

    # unpack
    results_dir = folderinfo["results_dir"]

    # initial message
    line_row = "-" * INFO_TEXT_WIDTH
    info_string = which_test + " - " + stats_var
    info_width = len(info_string)
    if info_width < INFO_TEXT_WIDTH:
        side_space = " " * ((INFO_TEXT_WIDTH - info_width) // 2)
    else:
        side_space = ""  # dont bother for extremely long features
    message = (
        "\n\n"
        + line_row
        + "\n\n"
        + side_space
        + info_string
        + side_space
        + "\n\n"
        + line_row
    )

    # print & save
    print(message)
    stats_textfile = os.path.join(results_dir, STATS_TXT_FILENAME)
    with open(stats_textfile, "a") as f:
        f.write(message)


def save_stats_summary_to_text(
    results_df, which_test, folderinfo, cfg, ANOVA_result=None
):
    """Save the numerical results of our cluster extent or ANOVA results to a text file
    Note
    ----
    which_test can either be:
        "RM ANOVA", "Mixed ANOVA" or "Permutation Test"
        If in the future you want to have some other test be mindful about the:
            if "ANOVA" in which_test lines!
    """
    # unpack
    contrasts = folderinfo["contrasts"]
    results_dir = folderinfo["results_dir"]
    bin_num = cfg["bin_num"]
    stats_threshold = cfg["stats_threshold"]

    # initialise the message variable
    message = ""

    # only for ANOVA - add the table above contrast loop
    if "ANOVA" in which_test:
        message = (
            message
            + "\n\n--------------------\nA N O V A  T A B L E"
            + "\n--------------------"
            + "\n"
            + str(ANOVA_result)
        )

    # contrast specific info
    for contrast in contrasts:
        # extract significant clusters
        # => works slightly different based on which statistical test we used, but in
        #    both cases it returns a list of lists of indices with range(bin_num)
        #   - which is why we can use rounded_sc_percentages as we do below
        if "ANOVA" in which_test:
            clusters = extract_multcomp_significance_clusters(
                results_df, contrast, stats_threshold
            )
        else:
            clusters = extract_all_clusters(results_df, contrast)
        # write message
        this_width = len("C O M P A R I S O N")
        if len(contrast) < this_width:
            side_spaces = " " * ((this_width - len(contrast)) // 2)
        else:
            side_spaces = ""  # dont bother if it's long
        message = (
            message
            + "\n\n-------------------\nC O M P A R I S O N\n"
            + side_spaces
            + contrast
            + side_spaces
            + "\n-------------------\n"
        )
        if len(clusters) == 0:  # no sig clusters were found!
            message = message + "No significant clusters!"
        else:
            rounded_sc_percentages = np.linspace(0, 100, bin_num).round(2)
            message = message + "Significant clusters at:"
            for cluster in clusters:
                message = (
                    message
                    + "\n\n"
                    + str(rounded_sc_percentages[cluster[0]])
                    + "-"
                    + str(rounded_sc_percentages[cluster[1]])
                    + "%"
                    + ", p values:\n\n"
                )
                # add pvals to message
                # => handle ANOVA and perm test differently since results_df different
                if "ANOVA" in which_test:
                    # important to include cluster-end bin here thus cluster[1]+1!
                    for i in range(cluster[0], cluster[1] + 1):
                        message = (
                            message
                            + "\n"
                            + str(rounded_sc_percentages[i])
                            + "% - p = "
                            + str(
                                round(
                                    results_df.loc[
                                        i,
                                        MULTCOMP_RESULT_P_IDENTIFIER
                                        + MULTCOMP_RESULT_SPLIT_STR
                                        + contrast,
                                    ],
                                    4,
                                )
                            )
                        )
                else:
                    # extract subset df of only this contrast because cluster variables
                    # idxs values correspond to 0:bin_num and results_df of perm test
                    # (this is different for anovas) has 0:bin_num*contrast_number!
                    this_contrast_results_df = results_df[
                        results_df[CONTRASTS_COL] == contrast
                    ]
                    cluster_p_colidx = this_contrast_results_df.columns.get_loc(
                        CLUSTER_P_COL
                    )
                    # also note that we only use cluster[0] to retrieve cluster's pval
                    # since the pval is constant across the whole cluster always
                    message = (
                        message
                        + "Cluster p value = "
                        + str(
                            this_contrast_results_df.iloc[cluster[0], cluster_p_colidx]
                        )
                    )

    # print & save
    print(message)
    stats_textfile = os.path.join(results_dir, STATS_TXT_FILENAME)
    with open(stats_textfile, "a") as f:
        f.write(message)


def save_multcomp_pvalues_to_excel(multcomp_df, stats_var, folderinfo, cfg):
    """Save all p-values of all contrasts to an excel file"""

    # unpack
    stats_threshold = cfg["stats_threshold"]
    contrasts = folderinfo["contrasts"]
    results_dir = folderinfo["results_dir"]

    #  initialise
    if os.path.exists(os.path.join(results_dir, MULTCOMP_EXCEL_FILENAME_1)):
        workbook = openpyxl.load_workbook(
            os.path.join(results_dir, MULTCOMP_EXCEL_FILENAME_1)
        )
        # create a new sheet for this var and make it active
        new_sheet = workbook.create_sheet(title=stats_var)
        sheet = new_sheet
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
    sheet.title = stats_var

    # this row idx used throughout excel table
    this_row_idx = 1
    row_height_per_sc = len(contrasts) + 2  # +2 for an 1) empty start row & 2) sc % row

    # add column headers
    # => add stats var manually before looping over cols 2:end using constant var
    cell = sheet["A" + str(this_row_idx)]
    cell.value = stats_var
    cell.font = openpyxl.styles.Font(bold=True)
    for c, col in enumerate(MULTCOMP_EXCEL_COLS):
        cell = sheet[string.ascii_uppercase[c + 1] + str(this_row_idx)]  # note the c+1!
        cell.value = col
        cell.font = openpyxl.styles.Font(bold=True)

    # add cell values: loop over sc percentage and contrasts and fill results correctly
    # from multcomp_df as you go
    # => make sure to use empty rows as well
    for sc_idx, sc_val in enumerate(multcomp_df[SC_PERCENTAGE_COL]):
        if sc_idx > 0:  # update row index of the excel file correctly
            this_row_idx += row_height_per_sc
        sheet.cell(row=this_row_idx + 1, column=1, value="")  # empty row
        sheet.cell(row=this_row_idx + 2, column=1, value=f"{sc_val}% cycle")  # sc % row
        for contrast_idx, contrast in enumerate(contrasts):
            # initialise this contrast's row index
            this_contrast_row_idx = this_row_idx + contrast_idx + 3
            # col 1: just a string of the contrast
            sheet.cell(
                row=this_contrast_row_idx,
                column=1,
                value=contrast,
            )
            # col 2: Tukey's q
            this_q = multcomp_df.loc[sc_idx, "q" + MULTCOMP_RESULT_SPLIT_STR + contrast]
            sheet.cell(
                row=this_contrast_row_idx,
                column=2,
                value=this_q,
            )
            # col 3: p value (rounded to 4 decimals)
            this_p = multcomp_df.loc[
                sc_idx, "p" + MULTCOMP_RESULT_SPLIT_STR + contrast
            ].round(4)
            sheet.cell(
                row=this_contrast_row_idx,
                column=3,
                value=this_p,
            )
            # col 4: Confidence Interval
            this_CI_list = []
            for CI_col_str in ["CI low", "CI high"]:
                this_CI_list.append(
                    multcomp_df.loc[
                        sc_idx, CI_col_str + MULTCOMP_RESULT_SPLIT_STR + contrast
                    ].round(4)
                )
            this_CI = str(this_CI_list[0]) + " to " + str(this_CI_list[1])
            sheet.cell(
                row=this_contrast_row_idx,
                column=4,
                value=this_CI,
            )
            # col 5: reject H0
            if this_p < stats_threshold:
                this_mask = "Yes"
            else:
                this_mask = "No"
            sheet.cell(
                row=this_contrast_row_idx,
                column=5,
                value=this_mask,
            )
            # col 6: significance level
            if this_p < 0.05:
                this_sig_level = "*"
            elif this_p < 0.01:
                this_sig_level = "**"
            elif this_p < 0.001:
                this_sig_level = "***"
            else:
                this_sig_level = "n.s."
            sheet.cell(
                row=this_contrast_row_idx,
                column=6,
                value=this_sig_level,
            )

    # save
    workbook.save(os.path.join(results_dir, MULTCOMP_EXCEL_FILENAME_1))

    # also save multcomp_df in case some people prefer that version
    multcomp_excel_file_2 = os.path.join(results_dir, MULTCOMP_EXCEL_FILENAME_2)
    if os.path.exists(multcomp_excel_file_2):
        with pd.ExcelWriter(multcomp_excel_file_2, mode="a") as writer:
            multcomp_df.to_excel(writer, sheet_name=stats_var, index=False)
    else:
        multcomp_df.to_excel(
            multcomp_excel_file_2,
            sheet_name=stats_var,
            index=False,
        )
