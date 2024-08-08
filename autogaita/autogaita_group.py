# %% imports
import os
import sys
import json
import string
import pandas as pd
import numpy as np
from sklearn.preprocessing import StandardScaler
from sklearn.decomposition import PCA
import openpyxl
from sklearn.utils import shuffle
import matplotlib
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
from matplotlib.animation import FuncAnimation, FFMpegWriter
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
import tkinter as tk
import customtkinter as ctk
from pingouin import sphericity, mixed_anova
from scipy import stats
import seaborn as sns
import warnings

# %% A note on cross species functionality
# => This function supports cross species analyses, however the data must be obtained
#    from the same tracking_software (=> I.e., resulting from autogaita_dlc or _simi)
# !!!
# Add "both" leg functionality in compute_avg & g_avg_dfs at some point. For now make
# sure it works for which_leg = right for our CS paper
# !!!
# The code won't work for tracking_software differing within a given execution of the
# script. Not sure the hassle will be worth it, particularly since it's pretty unlikely
# that a given lab has both DLC and Simi setups


# %% constants
# SET PLT BACKEND
# Agg is a non-interactive backend for plotting that can only write to files
# this is used to generate and save the plot figures
# later a tkinter backend (FigureCanvasTkAgg) is used for the plot panel
matplotlib.use("agg")

# increase resolution of figures
plt.rcParams["figure.dpi"] = 300

# Issues & stats results are stored in these textfiles (config json from mouseanalysis)
ISSUES_TXT_FILENAME = "Issues.txt"
STATS_TXT_FILENAME = "Stats.txt"
CONFIG_JSON_FILENAME = "config.json"

# EXPORT XLS
NORM_SHEET_NAME = "Normalised Stepcycles"
ORIG_SHEET_NAME = "Original Stepcycles"
NORM_GROUP_SHEET_NAME = "Normalised Group Stepcycles"
ORIG_GROUP_SHEET_NAME = "Original Group Stepcycles"
AVG_GROUP_SHEET_NAME = "Average Group Stepcycles"
STD_GROUP_SHEET_NAME = "Standard Deviation Group Stepcycles"
G_AVG_GROUP_SHEET_NAME = "Grand Average Group Stepcycles"
G_STD_GROUP_SHEET_NAME = "Grand Standard Deviation Group Stepcycles"

# SPLIT STRING (for _dlc first-level) & COLS OF DFs CREATED IN THIS SCRIPT
SPLIT_STRING = " - "
ID_COL = "ID"
SC_NUM_COL = "SC Number"
GROUP_COL = "Group"
N_COL = "N"  # for grand average dfs
SC_PERCENTAGE_COL = "SC Percentage"

# STATS
CONTRASTS_COL = "Contrasts"
TTEST_MASK_THRESHOLD = 0.05
TTEST_P_COL = "Ttest p"
TTEST_T_COL = "Ttest t"
TTEST_MASK_COL = "Ttest Mask"
CLUSTER_TMASS_COL = "Cluster Tmass"
CLUSTER_P_COL = "Cluster p"
CLUSTER_MASK_COL = "Cluster Mask"

# PLOTS
PERM_PLOT_LEGEND_SIZE = 6
PERM_PLOT_SUPLABEL_SIZE = 12
BOX_COLOR = "#fe420f"  # significance boxes - col = orangered
BOX_ALPHA = 0.1
STD_ALPHA = 0.2  # std boxes around means
STD_LW = 0

# PLOT GUI COLORS
FG_COLOR = "#5a7d9a"  # steel blue
HOVER_COLOR = "#8ab8fe"  # carolina blue


# %% main


def group(folderinfo, cfg):
    """Runs the main program for a group-level analysis comparing 2-5 groups

    Procedure
    ---------
    1) prepare some cfg and folderinfo - e.g., read bin_num/read jsons/global vars
    2) import the results folders and create dfs and raw_dfs (non-standardised) SC data
    3) compute average/std (ID-level) and grand-average/grand-std dataframes
    4) PCA
    5) prepare stats: create stats_df
    6) perform the cluster-extent permutation test
    7) perform the RM-/Mixed-ANOVA
    8) plots
    """
    # .............. initiate plot panel class and build loading screen ................
    # create class instance independently of "dont_show_plots" to not break the code
    plot_panel_instance = PlotPanel()

    if cfg["dont_show_plots"] is True:
        pass  # going on without building the loading screen

    elif cfg["dont_show_plots"] is False:  # -> show plot panel
        # build loading screen
        plot_panel_instance.build_plot_panel_loading_screen()

    # ..............................  print start  ....................................
    print_start(folderinfo, cfg)

    # ..................................  unpack  ......................................
    folderinfo, cfg = some_prep(folderinfo, cfg)

    # ...................................  import  .....................................
    dfs, raw_dfs, cfg = import_data(folderinfo, cfg)

    # .................................  avgs & stds  ..................................
    avg_dfs, std_dfs = avg_and_std(dfs, folderinfo, cfg)

    # ..............................  grand avgs & stds  ...............................
    g_avg_dfs, g_std_dfs = grand_avg_and_std(avg_dfs, folderinfo, cfg)

    # ...................................  PCA  ........................................
    if cfg["PCA_variables"]:  # empty lists are falsey!
        PCA_on_a_limb(avg_dfs, folderinfo, cfg, plot_panel_instance)
    plt.close("all")  # OK since all figures passed to save-funcs & PlotPanel

    # ..............................  prepare statistics  ..............................
    stats_df = create_stats_df(avg_dfs, folderinfo, cfg)

    # ......................  cluster-extent permutation test  .........................
    if cfg["stats_variables"]:  # empty lists are falsey!
        if cfg["do_permtest"]:
            for stats_var in cfg["stats_variables"]:
                cluster_extent_test(
                    stats_df,
                    g_avg_dfs,
                    g_std_dfs,
                    stats_var,
                    folderinfo,
                    cfg,
                    plot_panel_instance,
                )
        plt.close("all")

        # ..................................  ANOVA  ...................................
        if cfg["do_anova"]:  # indentation since we check for stats-vars here too!
            anova_sanity = anova_design_sanity_check(stats_df, folderinfo, cfg)
            if anova_sanity:
                for stats_var in cfg["stats_variables"]:
                    twoway_RMANOVA(
                        stats_df,
                        g_avg_dfs,
                        g_std_dfs,
                        stats_var,
                        folderinfo,
                        cfg,
                        plot_panel_instance,
                    )
        plt.close("all")

    # ..................................  plots  .......................................
    plot_results(g_avg_dfs, g_std_dfs, folderinfo, cfg, plot_panel_instance)

    # ..............................  print finish  ....................................
    print_finish(folderinfo)


# %% .................  local functions #0 - unpack & prepare vars  ....................
def some_prep(folderinfo, cfg):
    """Add some folderinfo & cfg variables to the dictionaries for further processes"""

    # unpack
    group_names = folderinfo["group_names"]
    group_dirs = folderinfo["group_dirs"]
    results_dir = folderinfo["results_dir"]

    # names & paths to folderinfo
    if not os.path.isdir(results_dir):
        os.makedirs(results_dir)
    folderinfo["contrasts"] = []
    for i in range(len(group_names)):
        for j in range(i + 1, len(group_names)):
            contrast = group_names[i] + " & " + group_names[j]
            folderinfo["contrasts"].append(contrast)

    # check if we previously had saved issues & stats textfiles, if so delete them
    issues_txt_path = os.path.join(results_dir, ISSUES_TXT_FILENAME)
    if os.path.exists(issues_txt_path):
        os.remove(issues_txt_path)
    stats_txt_path = os.path.join(results_dir, STATS_TXT_FILENAME)
    if os.path.exists(stats_txt_path):
        os.remove(stats_txt_path)

    # extracted_cfg_vars: save_to_xls, PCA tests & dont show plots
    cfg = extract_cfg_vars(folderinfo, cfg)

    # see if there's a config json file and add to cfg dict
    for g_idx, group_dir in enumerate(group_dirs):
        with open(
            os.path.join(group_dir, CONFIG_JSON_FILENAME), "r"
        ) as config_json_file:
            config_vars_from_json = json.load(config_json_file)
            for key in config_vars_from_json.keys():
                # assigning like this ensure all keys are in all jsons across groups
                if g_idx == 0:
                    cfg[key] = config_vars_from_json[key]
                else:
                    # sanity check for group-differences in cfg variables!
                    if (key not in cfg.keys()) | (
                        cfg[key] != config_vars_from_json[key]
                    ):
                        error_message = (
                            "config.json variables differ between groups!"
                            + "\nPlease make sure that all cfg variables between "
                            + "groups match & try again!"
                        )
                        raise ValueError(error_message)
                    else:
                        cfg[key] = config_vars_from_json[key]

    # after having checked cfg keys for equivalence, we have to make sure that
    # hind_joints is renamed to joints if DLC
    # => note that cfg will not be updated (and overwritten from input if Simi so it's)
    #    okay that we do this
    if cfg["tracking_software"] == "DLC":
        cfg["joints"] = cfg["hind_joints"]
    joints = cfg["joints"]
    angles = cfg["angles"]

    # prepare some plotting color stuff
    cfg["group_color_cycler"] = plt.cycler(
        "color", sns.color_palette(cfg["color_palette"], len(group_names))
    )
    cfg["group_color_dict"] = dict(
        zip(group_names, cfg["group_color_cycler"].by_key()["color"])
    )
    cfg["joint_color_cycler"] = plt.cycler(
        "color", sns.color_palette(cfg["color_palette"], len(joints))
    )
    cfg["angle_color_cycler"] = plt.cycler(
        "color", sns.color_palette(cfg["color_palette"], len(angles["name"]))
    )

    return folderinfo, cfg


def extract_cfg_vars(folderinfo, cfg):
    """Extract bin_num and save_to_xls from example Normalised dfs and sanity check
    that they match between groups!
    """

    group_names = folderinfo["group_names"]
    group_dirs = folderinfo["group_dirs"]
    results_dir = folderinfo["results_dir"]

    # ................................  save_to_xls  ...................................
    save_to_xls = [None] * len(group_dirs)
    for g, group_dir in enumerate(group_dirs):
        all_results_folders = os.listdir(
            group_dir
        )  # remove no-results valid_results_folders
        valid_results_folders = []
        # => Note if there's ambiguity / mixed filetypes, we set save_to_xls to True!
        sheet_type_mismatch_message = (
            "\n***********\n! WARNING !\n***********\n"
            + "Mismatch in sheet file types for group "
            + group_names[g]
            + "!\nSaving all output sheets to"
            + ".xlsx!\nRe-run first level & only save .csvs if "
            + "you want .csv files of group results!"
        )
        for folder in all_results_folders:
            # create save_to_xls here, there are two cases we have to deal with:
            # case 1: we found a csv file
            if os.path.exists(
                os.path.join(
                    group_dir,
                    folder,
                    folder + " - " + ORIG_SHEET_NAME + ".csv",
                )
            ):
                valid_results_folders.append(folder)
                if save_to_xls[g] is None:
                    save_to_xls[g] = False
                if save_to_xls[g] is True:
                    print(sheet_type_mismatch_message)
                    write_issues_to_textfile(sheet_type_mismatch_message, results_dir)
            # case 2: we found a xlsx file
            elif os.path.exists(
                os.path.join(
                    group_dir,
                    folder,
                    folder + " - " + ORIG_SHEET_NAME + ".xlsx",
                )
            ):
                valid_results_folders.append(folder)
                if save_to_xls[g] is None:
                    save_to_xls[g] = True
                if save_to_xls[g] is False:
                    save_to_xls[g] = True
                    print(sheet_type_mismatch_message)
                    write_issues_to_textfile(sheet_type_mismatch_message, results_dir)
        # test that at least 1 folder has valid results for all groups
        if not valid_results_folders:
            no_valid_results_error = (
                "\n*********\n! ERROR !\n*********\n"
                + "No valid results folder found for "
                + group_names[g]
                + "\nFix & re-run!"
            )
            print(no_valid_results_error)
            write_issues_to_textfile(no_valid_results_error, results_dir)
    # assign to our cfg dict after group loop
    cfg["save_to_xls"] = save_to_xls

    # .........................  test if PCA config is valid  ..........................
    if cfg["PCA_variables"]:  # only test if user wants PCA (ie. selected any features)
        if len(cfg["PCA_variables"]) < cfg["number_of_PCs"]:
            PCA_variable_num = len(cfg["PCA_variables"])
            PCA_variables_str = "\n".join(cfg["PCA_variables"])
            PCA_error_message = (
                "\n*********\n! ERROR !\n*********\n"
                + "\nPCA Configuration invalid, number of input features cannot exceed "
                + "number of principal components to compute!\n"
                + str(PCA_variable_num)
                + " PCA variables: \n"
                + PCA_variables_str
                + "\n & Number of wanted PCs: "
                + str(cfg["number_of_PCs"])
                + "\n Fix & re-run!"
            )
            write_issues_to_textfile(PCA_error_message, results_dir)
            raise ValueError(PCA_error_message)
        if cfg["number_of_PCs"] < 2:
            print(
                "\n***********\n! WARNING !\n***********\n"
                + "Number of principal components of PCA cannot be smaller than 2!"
                + "\nRunning PCA on 2 components - if you do not want to perform PCA, "
                + "just don't choose any variables for it."
            )
            cfg["number_of_PCs"] = 2  # make sure to update in cfg dict

    return cfg


# %% .................  local functions #1 - import run-level results  .................


# .................................  main functions  ...................................
def import_data(folderinfo, cfg):
    """Loop over all valid_results_folders of each group's /Results/ folder and create
    dfs of normalised and original (latter called "_raw") step-cycle datasets
    """

    # unpack
    group_names = folderinfo["group_names"]
    group_dirs = folderinfo["group_dirs"]
    results_dir = folderinfo["results_dir"]
    save_to_xls = cfg["save_to_xls"]
    which_leg = cfg["which_leg"]
    tracking_software = cfg["tracking_software"]

    # prepare lists of group-level dfs
    dfs = [pd.DataFrame(data=None)] * len(group_names)
    raw_dfs = [pd.DataFrame(data=None)] * len(group_names)
    # loop over each subfolder in each group-dir (i.e. "Results")
    for g, group_dir in enumerate(group_dirs):
        group_name = group_names[g]  # for import and combine function
        # valid_results_folders is the subset of all_results_folders in which valid
        # results were found (i.e., at least 1 valid SC was extracted & analysed)
        all_results_folders = os.listdir(group_dir)
        valid_results_folders = []
        for folder in all_results_folders:
            if (
                os.path.exists(
                    os.path.join(
                        group_dir,
                        folder,
                        folder + " - " + ORIG_SHEET_NAME + ".csv",
                    )
                )
            ) | (
                os.path.exists(
                    os.path.join(
                        group_dir,
                        folder,
                        folder + " - " + ORIG_SHEET_NAME + ".xlsx",
                    )
                )
            ):
                valid_results_folders.append(folder)
        # below local function keeps loading group-dfs and adds current dataset to it
        # => called once for normalised and once for original data
        for name in valid_results_folders:
            dfs[g] = import_and_combine_dfs(
                dfs[g],
                group_name,
                group_dir,
                tracking_software,
                name,
                "Normalised",
                which_leg,
                results_dir,
                cfg,
            )
            raw_dfs[g] = import_and_combine_dfs(
                raw_dfs[g],
                group_name,
                group_dir,
                tracking_software,
                name,
                "Original",
                which_leg,
                results_dir,
                cfg,
            )
        # reorder the columns we added
        if tracking_software == "DLC":
            cols = [ID_COL, "Run", "Stepcycle", "Flipped", "Time"]
        elif tracking_software == "Simi":
            cols = [ID_COL, "Leg", "Stepcycle", "Time"]
        dfs[g] = dfs[g][cols + [c for c in dfs[g].columns if c not in cols]]
        raw_dfs[g] = raw_dfs[g][cols + [c for c in raw_dfs[g].columns if c not in cols]]
        # for both dfs, check if there's rows with consecutive np.nan entries
        # => this happened while testing for some strange edge case I didn't understand)
        # => if so, remove them so we only have 1 row of np.nan
        all_nan_df = dfs[g].isna().all(axis=1)  # normalised dfs
        consecutive_nan_df = all_nan_df & all_nan_df.shift(fill_value=False)
        dfs[g] = dfs[g][~consecutive_nan_df]
        all_nan_raw_df = raw_dfs[g].isna().all(axis=1)  # original (raw) dfs
        consecutive_nan_raw_df = all_nan_raw_df & all_nan_raw_df.shift(fill_value=False)
        raw_dfs[g] = raw_dfs[g][~consecutive_nan_raw_df]
        # save to files & return
        norm_filepath = os.path.join(
            results_dir, group_names[g] + " - " + NORM_GROUP_SHEET_NAME  # norm SCs
        )
        save_results_sheet(dfs[g], save_to_xls[g], norm_filepath)
        orig_filepath = os.path.join(
            results_dir, group_names[g] + " - " + ORIG_GROUP_SHEET_NAME  # orig SCs
        )
        save_results_sheet(raw_dfs[g], save_to_xls[g], orig_filepath)
    # test: is bin_num is consistent across our groups
    # => if so, add it as well as one_bin_in_% to cfg
    cfg["bin_num"] = test_bin_num_consistency(dfs, group_names, results_dir)
    cfg["one_bin_in_sc_percent"] = 100 / cfg["bin_num"]
    return dfs, raw_dfs, cfg


def test_bin_num_consistency(dfs, group_names, results_dir):
    """Tests if bin number of step-cycle normalisation is consistent across groups"""
    # For this we use sc_breaks [i.e. the nan step-separators]) to see if bin_num is the
    # same value across all individual step-cycles across all groups.
    # Raise ValueError, stop everything & save info to Issues textfile if bin_num
    # should change at some point!
    # ==> check if normalised SC length is equal for all SCs
    # ==> break after SC1 == len of normalised SCs
    # ==> 0:24 = len of 25, with 25 being idx of first (nan) break

    bin_num = 0
    for g, group_name in enumerate(group_names):
        sc_breaks = np.where(pd.isnull(dfs[g][ID_COL]))[0]
        if bin_num == 0:
            bin_num = sc_breaks[0]
        for b in range(1, len(sc_breaks)):
            this_bin_num = sc_breaks[b] - sc_breaks[b - 1] - 1
            if this_bin_num != bin_num:
                bin_num_error_helper_function(
                    dfs, g, group_names, sc_breaks, results_dir, b
                )
        # handle the last step-cycle of the df (it doesn't have a sc_break after it!)
        if (len(dfs[g]) - sc_breaks[-1] - 1) != this_bin_num:
            bin_num_error_helper_function(
                dfs, g, group_names, sc_breaks, results_dir, b
            )
    return bin_num


def bin_num_error_helper_function(dfs, g, group_names, sc_breaks, results_dir, b):
    """Handle this error in a separate function for readability"""
    id_col_idx = dfs[g].columns.get_loc(ID_COL)
    message = (
        "\n*********\n! ERROR !\n*********\n"
        + "\nSC Normalisation bin number mismatch for:"
        + "\nGroup: "
        + group_names[g]
        + " - ID: "
        + str(dfs[g].iloc[sc_breaks[b] - 3, id_col_idx])
        + "\nPlease re-run & make sure all bin numbers match!"
    )
    print(message)
    write_issues_to_textfile(message, results_dir)
    raise ValueError(message)


def test_PCA_and_stats_variables(df, group_name, name, results_dir, cfg):
    """Tests if PCA & stats variables are present in all groups' datasets"""
    PCA_variables = cfg["PCA_variables"]
    stats_variables = cfg["stats_variables"]
    if PCA_variables:
        if not all(variable in df.columns for variable in PCA_variables):
            missing_PCA_variables = [
                variable for variable in PCA_variables if variable not in df.columns
            ]
            missing_PCA_variables_str = "\n".join(missing_PCA_variables)
            PCA_variable_mismatch_message = (
                "\n*********\n! ERROR !\n*********\n"
                + "\nNot all features you asked us to analyse WITH PCA were present in "
                + "the dataset of:\n"
                + group_name
                + "'s "
                + name
                + "\nMissing variables were:\n"
                + missing_PCA_variables_str
                + "\nPlease re-run & make sure all variables are present!"
            )
            print(PCA_variable_mismatch_message)
            write_issues_to_textfile(PCA_variable_mismatch_message, results_dir)
            raise ValueError(PCA_variable_mismatch_message)
    if stats_variables:
        if not all(variable in df.columns for variable in stats_variables):
            missing_stats_variables = [
                variable for variable in stats_variables if variable not in df.columns
            ]
            missing_stats_variables_str = "\n".join(missing_stats_variables)
            stats_variable_mismatch_message = (
                "\n*********\n! ERROR !\n*********\n"
                + "\nNot all features you asked us to analyse STATISTICALLY were "
                + "present in the dataset of:\n"
                + group_name
                + "'s "
                + name
                + "\nMissing variables were:\n"
                + missing_stats_variables_str
                + "\nPlease re-run & make sure all variables are present!"
            )
            print(stats_variable_mismatch_message)
            write_issues_to_textfile(stats_variable_mismatch_message, results_dir)
            raise ValueError(stats_variable_mismatch_message)


def import_and_combine_dfs(
    group_df,
    group_name,
    group_dir,
    tracking_software,
    name,
    which_df,
    which_leg,
    results_dir,
    cfg,
):
    """Import one run's df at a time and combine to group-level df"""
    if which_df == "Normalised":
        this_sheet_name = NORM_SHEET_NAME
    elif which_df == "Original":
        this_sheet_name = ORIG_SHEET_NAME
    fullfilepath = os.path.join(group_dir, name, name + " - " + this_sheet_name)
    if tracking_software == "DLC":
        df = load_sheet_file(fullfilepath)
    elif tracking_software == "Simi":
        df = load_sheet_file(fullfilepath, which_leg=which_leg)
    if df is None:
        this_message = (
            "\n***********\n! WARNING !\n***********\n"
            + "No "
            + which_df
            + " results sheet file found for "
            + name
            + " at\n"
            + group_dir
            + ".\nSkipping!"
        )
        print(this_message)
        write_issues_to_textfile(this_message, results_dir)
    else:
        df_copy = df.copy()
        # test: are our PCA & stats variables present in this ID's dataset?
        test_PCA_and_stats_variables(df_copy, group_name, name, results_dir, cfg)
        # add some final info & append to group_df
        if tracking_software == "DLC":
            # add this run's info to df (& prepare Stepcycle column)
            # => I call DLC stuff ID NUM - RUN NUM, so we can use temp_split & idxing
            #    as done below
            temp_split = name.split(SPLIT_STRING)
            mouse_num = int(temp_split[0].split(" ")[1])  # temp_split[0] == ID X
            run_num = int(temp_split[1].split(" ")[1])  # temp_split[1] == RUN X
            df_copy["Run"] = run_num
            df_copy[ID_COL] = mouse_num
        elif tracking_software == "Simi":
            df_copy[ID_COL] = name
        df_copy["Stepcycle"] = np.nan
        sc_col_idx = df_copy.columns.get_loc("Stepcycle")
        # add this df to group dfs
        sc_idxs = extract_sc_idxs(df_copy)
        for sc in range(len(sc_idxs)):
            df_copy.iloc[sc_idxs[sc], sc_col_idx] = sc + 1  # stepcycle info column
            if group_df.empty:
                group_df = df_copy.iloc[sc_idxs[sc]]
            else:
                nanvector = df_copy.loc[[1]]
                # new for pandas 2.2.2 - change all cols to floats so its compatible
                # with np.nan concatenation (does this break anything?)
                # => 07.08. - Leaving this for when I have unit tests for simi & group
                # => Not sure it'll work well like this I feel like it will take more
                #    than just changing the dtype
                # => For now ignoring warnings so users don't get annoyed (and because
                #    dtype incompatibility doesn't matter too much for us)
                # nanvector = nanvector.astype(float)
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore")
                    nanvector[:] = np.nan
                group_df = pd.concat([group_df, nanvector], axis=0)
                group_df = pd.concat([group_df, df_copy.iloc[sc_idxs[sc]]], axis=0)
    # if there was no valid sheet file for this name, we are returning group_df without
    # it - we warned the user
    return group_df


# ................................  helper functions  ..................................


def save_results_sheet(dataframe, save_to_xls, fullfilepath):
    """Save a csv or xls of results"""
    if save_to_xls:
        dataframe.to_excel(fullfilepath + ".xlsx", index=False)
    else:
        dataframe.to_csv(fullfilepath + ".csv", index=False)


def load_sheet_file(fullfilepath, **kwargs):
    """Handle that the user might have saved csv or xls files

    Important
    ---------
    ==> We also handle that we have 3 sheets for human analyses, user has to decide
        which leg (sheet) they want to analyse
    ==> We keep this independent of save_to_xls, because save_to_xls allows us to save
        group outputs to xlsx if some IDs were saved to csv and others to xlsx.. But,
        here we have to be sure that we can load both cases, irrespective of what
        save_to_xls is!
    """
    try:  # see if we have an exel file
        if "which_leg" in kwargs:
            which_leg = kwargs["which_leg"]
            df = pd.read_excel(fullfilepath + ".xlsx", sheet_name=which_leg)
        else:
            df = pd.read_excel(fullfilepath + ".xlsx")
        return df
    except FileNotFoundError:
        try:  # see if we have a csv file
            df = pd.read_csv(fullfilepath + ".csv")
            return df
        # if both are not found there are no files
        # => note that this is unlikely since we previously made sure to only
        #    consider valid_results_folders that had a normalised sheet file
        except FileNotFoundError:
            return


def extract_sc_idxs(df):
    """Extract step-cycle indices using original & normalised CSV files"""
    # A note on df & nan (see xls_separations):
    # ==> (Using range & iloc so we don't have to subtract 1 to nan-idxs)
    # ==> if there is more than 1 SC found, the first row of nan indicates the
    #     END of SC 1
    # ==> the last row of nan indicates the START of the last SC
    # ==> everything inbetween is alternatingly: (if you add 1 to nan-idx) the
    #     start of an SC + (if you subtract -1 to nan-idx) the end of that SC
    # ==> E.g.: separations[1]+1 is 1st idx of SC2 - separations[2]-1 is last
    #     idx of SC2
    first_col = df.columns[0]
    xls_separations = np.where(pd.isnull(df[first_col]))[0]
    sc_idxs = []
    # the next line avgs that we have exactly one step, because we would not
    # build df (and have results in the first place) if there was no step
    # Thus, if xls_sep. is empty (len=0) it avgs that no separations were
    # there, i.e., 1 SC
    if len(xls_separations) == 0:
        sc_idxs.append(range(0, len(df)))  # I can do this bc. only 1 SC
    else:
        for b in range(len(xls_separations)):
            if b == 0:
                # SC1 - 0 to (not including) nan/end-idx
                sc_idxs.append(range(xls_separations[b]))
            elif b > 0:  # inbetween SCs
                if (b % 2) == 0:
                    sc_idxs.append(
                        range(
                            xls_separations[b - 1] + 1,  # add 1=start
                            xls_separations[b],
                        )
                    )
            # last SC - I can write it this way because b will always be odd if
            # it refers to the start of a stepcycle & thus: possibly the last)
            if xls_separations[b] == xls_separations[-1]:
                sc_idxs.append(range(xls_separations[-1] + 1, len(df)))
    return sc_idxs


# %% ................  local functions #2 - run-level averages & stds  .................


def avg_and_std(dfs, folderinfo, cfg):
    """Compute the avgs & standard deviations for all columns of df"""

    # !!! NU - This function can be significantly optimised. I can initialise
    # colstoexclude, the avg/std dfs as a list of len=2, loop over them, and fix the
    # below NU.
    # => Keep this in mind but for now make sure it works for human data

    # unpack
    group_names = folderinfo["group_names"]
    results_dir = folderinfo["results_dir"]
    one_bin_in_sc_percent = cfg["one_bin_in_sc_percent"]
    save_to_xls = cfg["save_to_xls"]
    tracking_software = cfg["tracking_software"]
    if tracking_software == "DLC":
        analyse_average_x = cfg["analyse_average_x"]
    elif tracking_software == "Simi":
        analyse_average_y = cfg["analyse_average_y"]

    # preparation, initialise avg/std dfs & colnames
    avg_dfs = [pd.DataFrame(data=None)] * len(group_names)
    std_dfs = [pd.DataFrame(data=None)] * len(group_names)

    # loop over all groups' dfs
    for g, group_name in enumerate(group_names):
        # extract this df and IDs
        if tracking_software == "DLC":
            cols_to_exclude = [ID_COL, "Run", "Stepcycle", "Flipped", "Time"]
            for col in dfs[g].columns:
                if col.endswith("likelihood"):
                    cols_to_exclude.append(col)
                if analyse_average_x is False:
                    if col.endswith(" x"):
                        cols_to_exclude.append(col)
        elif tracking_software == "Simi":
            cols_to_exclude = [ID_COL, "Leg", "Stepcycle", "Time"]
            for col in dfs[g].columns:
                if col.endswith("X"):
                    cols_to_exclude.append(col)
                if analyse_average_y is False:
                    if col.endswith("Y"):
                        cols_to_exclude.append(col)
        this_df = dfs[g]
        IDs = pd.unique(this_df[ID_COL])
        IDs = IDs[~pd.isnull(IDs)]  # get rid of nan separator values
        # loop over all IDs, create an avg & std df for each, concat to groupdf
        for ID in IDs:
            # ID avg and std dfs
            this_ID_avg_df = pd.DataFrame(data=None)
            this_ID_std_df = pd.DataFrame(data=None)
            # indices of this IDs SCs
            this_idxs_list = np.where(this_df[ID_COL] == ID)[0]
            this_SC_idxs = idxs_list_to_list_of_SC_lists(this_idxs_list)
            SC_num = len(this_SC_idxs)
            # loop over all columns that are not ID, compute avg & std dfs
            # of current ID
            for c, col in enumerate(this_df.columns):
                if col not in cols_to_exclude:
                    this_data = np.zeros([len(this_SC_idxs[0]), SC_num])
                    for SC, SC_idx in enumerate(this_SC_idxs):
                        # NOTE - because SC_idx is a list of integers, iloc
                        #        includes all of the indices values. This is
                        #        different to e.g. cases of slicing or using
                        #        ranges. We can thus just forward and use
                        #        the indices found previously as they are in
                        #        the list!
                        this_data[:, SC] = this_df.iloc[SC_idx, c]
                    this_avg = np.mean(this_data, 1)
                    std = np.std(this_data, 1)
                    this_ID_avg_df[col] = this_avg
                    this_ID_std_df[col] = std
            # add ID col
            this_ID_avg_df[ID_COL] = ID
            this_ID_std_df[ID_COL] = ID
            # add SC number col
            this_ID_avg_df[SC_NUM_COL] = SC_num
            this_ID_std_df[SC_NUM_COL] = SC_num
            # SC Percentage column
            this_ID_avg_df[SC_PERCENTAGE_COL] = 0
            sc_percentage_col_idx = this_ID_avg_df.columns.get_loc(SC_PERCENTAGE_COL)
            for i in range(len(this_ID_avg_df)):
                this_index = this_ID_avg_df.index[i]
                this_ID_avg_df.iloc[i, sc_percentage_col_idx] = (
                    one_bin_in_sc_percent + (this_index * one_bin_in_sc_percent)
                )
            this_ID_std_df[SC_PERCENTAGE_COL] = 0
            sc_percentage_col_idx = this_ID_std_df.columns.get_loc(SC_PERCENTAGE_COL)
            for i in range(len(this_ID_std_df)):
                this_index = this_ID_std_df.index[i]
                this_ID_std_df.iloc[i, sc_percentage_col_idx] = (
                    one_bin_in_sc_percent + (this_index * one_bin_in_sc_percent)
                )
            # add ID-level avg & std dfs to group-level dfs
            if avg_dfs[g].empty:
                avg_dfs[g] = this_ID_avg_df
                std_dfs[g] = this_ID_std_df
            else:
                avg_dfs[g] = pd.concat([avg_dfs[g], this_ID_avg_df], axis=0)
                std_dfs[g] = pd.concat([std_dfs[g], this_ID_std_df], axis=0)

        # !!! NU - fix this:
        # I'm sure there are better ways of doing this, but we put in ID_COL
        # as first col like this.. Best practice would probably be to
        # initialise these dfs with that column but I'm using .empty above and
        # I also didn't want to introduce bin_num just for this
        first_cols = [ID_COL, SC_NUM_COL, SC_PERCENTAGE_COL]
        avg_dfs[g] = avg_dfs[g][
            first_cols + [c for c in avg_dfs[g].columns if c not in first_cols]
        ]
        std_dfs[g] = std_dfs[g][
            first_cols + [c for c in std_dfs[g].columns if c not in first_cols]
        ]
        # export sheets
        avg_filepath = os.path.join(
            results_dir, group_names[g] + " - " + AVG_GROUP_SHEET_NAME  # av SCs
        )
        save_results_sheet(avg_dfs[g], save_to_xls, avg_filepath)
        std_filepath = os.path.join(
            results_dir, group_names[g] + " - " + STD_GROUP_SHEET_NAME  # std SCs
        )
        save_results_sheet(std_dfs[g], save_to_xls, std_filepath)
    return avg_dfs, std_dfs


def idxs_list_to_list_of_SC_lists(idxs):
    """Convert all idxs that have this ID to a list of ranges with individual
    SC's values being separate entries in list
    """
    SC_idxs = []
    i = 0
    this_sc_idxs = []
    for i in range(1, len(idxs)):
        if idxs[i - 1] == (idxs[i] - 1):
            this_sc_idxs.append(idxs[i - 1])
        else:
            # if we have a SC change, add the -1 idx (last idx of prev SC),
            # then append & reset
            this_sc_idxs.append(idxs[i - 1])
            SC_idxs.append(this_sc_idxs)
            this_sc_idxs = []
        # special case: if i is last element (len-1), we in previous iteration
        #               did add i-1 but we have to make sure to add the last
        #               idx, too (since we can't loop until len+1)
        if i == (len(idxs) - 1):
            this_sc_idxs.append(idxs[i])
            SC_idxs.append(this_sc_idxs)
    return SC_idxs


# %% ..................  local functions #3 - grand averages & stds  ...................


def grand_avg_and_std(avg_dfs, folderinfo, cfg):
    """Compute the grand averages & standard deviations for all columns of df"""

    # unpack
    group_names = folderinfo["group_names"]
    results_dir = folderinfo["results_dir"]
    bin_num = cfg["bin_num"]
    one_bin_in_sc_percent = cfg["one_bin_in_sc_percent"]
    save_to_xls = cfg["save_to_xls"]

    # preparation, initialise g_avg/std dfs
    g_avg_dfs = [pd.DataFrame(data=None)] * len(group_names)
    g_std_dfs = [pd.DataFrame(data=None)] * len(group_names)
    # loop over all groups' dfs
    for g, group_name in enumerate(group_names):
        # extract this df and IDs
        this_df = avg_dfs[g]
        IDs = np.unique(this_df[ID_COL])
        ID_num = len(IDs)  # also: for an "N" column later
        for c, col in enumerate(this_df.columns):
            if col != ID_COL:
                this_data = np.zeros([bin_num, ID_num])
                for i, ID in enumerate(IDs):
                    this_idxs = np.where(this_df[ID_COL] == ID)[0]
                    this_data[:, i] = this_df.iloc[this_idxs, c]
                # assign to grand av dfs
                g_avg_dfs[g][col] = np.mean(this_data, 1)
                g_std_dfs[g][col] = np.std(this_data, 1)
        # add the number of IDs that went into these grand averages and have
        # this be first column & add SC Percentage col and have it be 2nd col
        g_avg_dfs[g][N_COL] = ID_num  # N column
        g_std_dfs[g][N_COL] = ID_num
        # SC Percentage column
        g_avg_dfs[g][SC_PERCENTAGE_COL] = 0
        sc_percentage_col_idx = g_avg_dfs[g].columns.get_loc(SC_PERCENTAGE_COL)
        for i in range(len(g_avg_dfs[g])):
            this_index = g_avg_dfs[g].index[i]
            g_avg_dfs[g].iloc[i, sc_percentage_col_idx] = one_bin_in_sc_percent + (
                this_index * one_bin_in_sc_percent
            )
        g_std_dfs[g][SC_PERCENTAGE_COL] = 0
        sc_percentage_col_idx = g_std_dfs[g].columns.get_loc(SC_PERCENTAGE_COL)
        for i in range(len(g_std_dfs[g])):
            this_index = g_std_dfs[g].index[i]
            g_std_dfs[g].iloc[i, sc_percentage_col_idx] = one_bin_in_sc_percent + (
                this_index * one_bin_in_sc_percent
            )
        # reorder columns (N & SC% first)
        first_cols = [N_COL, SC_PERCENTAGE_COL]
        g_avg_dfs[g] = g_avg_dfs[g][
            first_cols + [c for c in g_avg_dfs[g].columns if c not in first_cols]
        ]
        g_std_dfs[g] = g_std_dfs[g][
            first_cols + [c for c in g_std_dfs[g].columns if c not in first_cols]
        ]
        # export sheets
        g_avg_filepath = os.path.join(
            results_dir, group_names[g] + " - " + G_AVG_GROUP_SHEET_NAME  # g_av SCs
        )
        save_results_sheet(g_avg_dfs[g], save_to_xls, g_avg_filepath)
        g_std_filepath = os.path.join(
            results_dir, group_names[g] + " - " + G_STD_GROUP_SHEET_NAME  # g_std SCs
        )
        save_results_sheet(g_std_dfs[g], save_to_xls, g_std_filepath)
    return g_avg_dfs, g_std_dfs


# %% ...........................  local functions #4 - PCA  ............................


def PCA_on_a_limb(avg_dfs, folderinfo, cfg, plot_panel_instance):
    """PCA on joint y values of a limb (mouse: hindlimb, humans: leg of interest)"""

    # print info
    print("\n*************** Computing PCA ***************\n")
    # create the input dataframe
    PCA_df, features = create_PCA_df(avg_dfs, folderinfo, cfg)
    # run the PCA
    PCA_df, PCA_info = run_PCA(PCA_df, features, cfg)
    # save PCA info to xlsx file
    PCA_info_to_xlsx(PCA_df, PCA_info, folderinfo, cfg)
    # plot the scatterplot
    plot_PCA(PCA_df, PCA_info, folderinfo, cfg, plot_panel_instance)


def PCA_info_to_xlsx(PCA_df, PCA_info, folderinfo, cfg):
    """Save the explained_var & eigenvectors of PCs to an xlsx file"""

    # unpack
    results_dir = folderinfo["results_dir"]
    number_of_PCs = cfg["number_of_PCs"]

    # initialise
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "PCA Info"

    # add column headers
    for pc in range(number_of_PCs):
        sheet[string.ascii_uppercase[pc + 1] + "1"] = "PC " + str(pc + 1)
    # add cell values: explained variance
    sheet.cell(row=2, column=1, value="Explained Var. (%)")
    for pc in range(number_of_PCs):
        sheet.cell(
            row=2, column=pc + 2, value=round(PCA_info["explained_vars"][pc] * 100, 2)
        )
    # add cell values: eigenvectors
    sheet.cell(row=4, column=1, value="Features")
    features = PCA_df.columns[2:-number_of_PCs].values
    for i in range(len(features)):
        # row i + 5 because excel starts counting at 1 and: row2=var_exp/row3=empty/
        # row5=feature heading
        sheet.cell(row=i + 5, column=1, value=features[i])
        for pc in range(number_of_PCs):
            # column is pc+2 because we want pc=0 to be in xlsx column 2 etc.
            sheet.cell(row=i + 5, column=pc + 2, value=PCA_info["components"][pc, i])
    # save
    workbook.save(os.path.join(results_dir, "PCA Info.xlsx"))


def create_PCA_df(avg_dfs, folderinfo, cfg):
    """Create a ID x ID_COL + features dataframe to be used by PCA"""

    # unpack
    group_names = folderinfo["group_names"]
    PCA_vars = cfg["PCA_variables"]
    bin_num = cfg["bin_num"]

    PCA_df = pd.DataFrame(data=None)
    # create a list of features for series & dfs (features are vars @ each SC % bin)
    features = []
    for var in PCA_vars:
        for b in range(bin_num):
            bin_in_percent = int(((1 + b) / bin_num) * 100)
            features.append(var + " " + str(bin_in_percent))
    # for each mouse, create a series to concat to PCA_df
    for g, group_name in enumerate(group_names):
        for ID in pd.unique(avg_dfs[g][ID_COL]):
            this_list = [group_name, ID]
            ID_row_idx = np.where(avg_dfs[g][ID_COL] == ID)[0]
            for var in PCA_vars:
                joint_col_idx = avg_dfs[g].columns.get_loc(var)
                # get data of this ID x joint combo
                this_data = list(avg_dfs[g].iloc[ID_row_idx, joint_col_idx].values)
                this_list.extend(this_data)  # so we have ID_COL as 1st value
            this_series = pd.Series(this_list)
            if PCA_df.empty:
                PCA_df = pd.DataFrame(this_series).transpose()
            else:
                # transpose series, transform to df and concat to row-axis
                PCA_df = pd.concat([PCA_df, this_series.to_frame().T], axis=0)
    # add colnames after the last mouse (makes concat'ing series 2 df easier)
    PCA_df.columns = [GROUP_COL] + [ID_COL] + features
    return PCA_df, features


def run_PCA(PCA_df, features, cfg):
    """Runs the PCA on a limb's feature (e.g. y or z coordinates)"""

    # unpack
    number_of_PCs = cfg["number_of_PCs"]

    # run
    PCA_model = PCA(n_components=number_of_PCs)
    x = PCA_df.loc[:, features].values
    # standardise here so that EACH FEATURE has mean=0 & std=1
    # => you can check this with np.mean/std
    # => you can also check that this operates on columns (features) and not rows (IDs)
    x = StandardScaler().fit_transform(x)
    PCs = PCA_model.fit_transform(x)
    for i in range(number_of_PCs):
        PCA_df["PC " + str(i + 1)] = PCs[:, i]
    PCA_info = {
        "explained_vars": PCA_model.explained_variance_ratio_,
        "components": PCA_model.components_,
    }
    return PCA_df, PCA_info


def plot_PCA(PCA_df, PCA_info, folderinfo, cfg, plot_panel_instance):
    """Plot a scatterplot and colour based on group name"""

    # unpack
    group_names = folderinfo["group_names"]
    results_dir = folderinfo["results_dir"]
    group_color_dict = cfg["group_color_dict"]
    number_of_PCs = cfg["number_of_PCs"]
    save_3D_PCA_video = cfg["save_3D_PCA_video"]
    dont_show_plots = cfg["dont_show_plots"]
    legend_outside = cfg["legend_outside"]

    # loop over groups, and call scatter 3x so we can have a correct legend
    f, ax = plt.subplots(1, 1)
    if number_of_PCs > 2:
        f_3d = plt.figure()
        ax_3d = f_3d.add_subplot(111, projection="3d")
    PC1_col_idx = PCA_df.columns.get_loc("PC 1")
    PC2_col_idx = PCA_df.columns.get_loc("PC 2")
    if number_of_PCs > 2:
        PC3_col_idx = PCA_df.columns.get_loc("PC 3")
    for g, group_name in enumerate(group_names):
        row_idxs = np.where(PCA_df[GROUP_COL] == group_name)[0]
        x = PCA_df.iloc[row_idxs, PC1_col_idx].values
        y = PCA_df.iloc[row_idxs, PC2_col_idx].values
        ax.scatter(x, y, color=group_color_dict[group_name], label=group_name)
        if number_of_PCs > 2:
            z = PCA_df.iloc[row_idxs, PC3_col_idx].values
            ax_3d.scatter(
                x, y, z, color=group_color_dict[group_name], s=60, label=group_name
            )
    # legend adjustments
    if legend_outside is True:
        ax.legend(group_names, loc="center left", bbox_to_anchor=(1, 0.5))
    elif legend_outside is False:
        ax.legend(group_names)
    ax.set_xlabel("PC 1")
    ax.set_ylabel("PC 2")
    ax.set_title(
        "Explained vars.: PC 1 - "
        + str(round(PCA_info["explained_vars"][0] * 100, 2))
        + "% | PC 2 -  "
        + str(round(PCA_info["explained_vars"][1] * 100, 2))
        + "%"
    )
    if number_of_PCs > 2:  # 3d scatterplot
        ax_3d.view_init(30, 125)
        # legend adjustments
        if legend_outside is True:
            ax_3d.legend(group_names, loc="center right", bbox_to_anchor=(0, 0.5))
        elif legend_outside is False:
            ax_3d.legend(group_names)
        ax_3d.set_xlabel("PC 1")
        ax_3d.set_ylabel("PC 2")
        ax_3d.set_zlabel("PC 3")
        ax_3d.set_title(
            "Explained vars.: PC 1 - "
            + str(round(PCA_info["explained_vars"][0] * 100, 2))
            + "% | PC 2 -  "
            + str(round(PCA_info["explained_vars"][1] * 100, 2))
            + "%"
            + "% | PC 3 -  "
            + str(round(PCA_info["explained_vars"][2] * 100, 2))
            + "%"
        )
    save_figures(f, results_dir, "PCA Scatterplot")
    # add figure to plot panel figures list
    if dont_show_plots is False:  # -> show plot panel
        plot_panel_instance.figures.append(f)

    # 3d scatterplot image file
    if number_of_PCs > 2:
        save_figures(f_3d, results_dir, "PCA 3D Scatterplot")

        # add figure to plot panel figures list
        if dont_show_plots is False:  # -> show plot panel
            plot_panel_instance.figures.append(f)

        # 3d scatterplot rotating video file
        if save_3D_PCA_video:

            print("*************** Saving 3D PCA Scatterplot ***************")

            def init():  # create the animation
                return (f_3d,)

            def animate(frame):
                ax_3d.view_init(elev=10, azim=frame)
                return (f_3d,)

            anim = FuncAnimation(
                f_3d, animate, frames=np.arange(0, 360, 1), interval=20, blit=True
            )
            writervideo = FFMpegWriter(fps=30)  # save to m4 using ffmpeg writer
            anim.save(results_dir + "PCA 3D Scatterplot.mp4", writer=writervideo)


# %% ................  local functions #5 - prepare statistics  ........................


def create_stats_df(avg_dfs, folderinfo, cfg):
    """Create a df in format used by (both) our stats approaches.
    Shape: ID*SC% x features (i.e. all IDs' averages concatenated along rows)
    """

    # unpack
    group_names = folderinfo["group_names"]
    one_bin_in_sc_percent = cfg["one_bin_in_sc_percent"]

    for g, group_name in enumerate(group_names):
        avg_dfs[g][GROUP_COL] = group_name
        if g == 0:
            stats_df = avg_dfs[g]
        else:
            stats_df = pd.concat([stats_df, avg_dfs[g]], axis=0)
    stats_df[SC_PERCENTAGE_COL] = 0
    sc_percentage_col_idx = stats_df.columns.get_loc(SC_PERCENTAGE_COL)
    for i in range(len(stats_df)):
        this_index = stats_df.index[i]
        stats_df.iloc[i, sc_percentage_col_idx] = one_bin_in_sc_percent + (
            this_index * one_bin_in_sc_percent
        )
    return stats_df


# %% ...........  local functions #6 - cluster-extent permutation test  ................


# ...............................  main function  ......................................
def cluster_extent_test(
    stats_df, g_avg_dfs, g_std_dfs, stats_var, folderinfo, cfg, plot_panel_instance
):
    """Main function running a cluster-extent permutation test of N contrasts for a
    given dependent variable
    """

    # unpack
    permutation_number = cfg["permutation_number"]
    stats_threshold = cfg["stats_threshold"]
    # true observed
    trueobs_df = stats_df.copy()
    trueobs_results_df = initialise_results_df(folderinfo, cfg)
    trueobs_results_df = compute_first_level_results(
        trueobs_df, trueobs_results_df, stats_var, folderinfo
    )
    # permutation
    max_tmass = np.zeros(permutation_number)  # tmass
    for p in range(permutation_number):
        permuted_df = permute_true_observed_df(trueobs_df, cfg)
        permuted_results_df = initialise_results_df(folderinfo, cfg)
        permuted_results_df = compute_first_level_results(
            permuted_df, permuted_results_df, stats_var, folderinfo
        )
        # max tmass
        max_tmass[p] = max(permuted_results_df[CLUSTER_TMASS_COL])
        sys.stdout.write(
            "\r*************** Permuting "
            + stats_var
            + ": "
            + str(p + 1)
            + "/"
            + str(permutation_number)
            + " ***************"
        )
        sys.stdout.flush()
    # assign final p values of true observed cluster sizes
    trueobs_results_df = test_trueobs_clusters(
        trueobs_results_df, max_tmass, permutation_number, stats_threshold
    )
    # print & save exact numerical results (significant SC % clusters) to a textfile
    save_stats_results_to_text(
        trueobs_results_df,
        stats_var,
        "Cluster-extent permutation test",
        folderinfo,
        cfg,
    )
    # plot results
    plot_permutation_test_results(
        g_avg_dfs,
        g_std_dfs,
        trueobs_results_df,
        stats_var,
        folderinfo,
        cfg,
        plot_panel_instance,
    )


# ................................  preparation  .......................................
def initialise_results_df(folderinfo, cfg):
    """Initialise a results df for permutation test."""

    # unpack
    contrasts = folderinfo["contrasts"]
    bin_num = cfg["bin_num"]

    initial_contrasts_list = []
    for contrast in contrasts:
        initial_contrasts_list.extend([contrast] * bin_num)
    results_df = pd.DataFrame(data=initial_contrasts_list, columns=[CONTRASTS_COL])
    results_df[SC_PERCENTAGE_COL] = None
    results_df[TTEST_P_COL] = float(1)
    results_df[TTEST_T_COL] = float(0)
    results_df[TTEST_MASK_COL] = False
    results_df[CLUSTER_TMASS_COL] = 0.0
    return results_df


# ...............................  first-level  ........................................
def compute_first_level_results(stats_df, results_df, stats_var, folderinfo):
    """Compute the results of our first level, i.e., mass ttests"""

    # unpack
    contrasts = folderinfo["contrasts"]

    # populate true observed results df
    idx = 0
    for contrast in contrasts:
        group1 = contrast.split(" & ")[0]
        group2 = contrast.split(" & ")[1]
        # SC percentage & ttest results
        for s, sc_percentage in enumerate(np.unique(stats_df[SC_PERCENTAGE_COL])):
            results_df = run_and_assign_ttest(
                stats_df,
                results_df,
                stats_var,
                contrast,
                group1,
                group2,
                sc_percentage,
                idx,
            )
            # update idx for assigning sc_perc in next iter correctly (function input)
            idx += 1
        # cluster size
        results_df = compute_and_assign_clustersize(results_df, contrast)
    return results_df


def run_and_assign_ttest(
    stats_df, results_df, stats_var, contrast, group1, group2, sc_percentage, idx
):
    """Run ttest for a given pair of groups & a given percentage of the step cycle."""
    # get location of current SC Percentage
    sc_percentage_col_idx = results_df.columns.get_loc(SC_PERCENTAGE_COL)
    results_df.iloc[idx, sc_percentage_col_idx] = sc_percentage
    # extract the two arrays to be tested, test & get its results
    arr1 = extract_variable_array(stats_df, stats_var, group1, sc_percentage)
    arr2 = extract_variable_array(stats_df, stats_var, group2, sc_percentage)
    this_t, this_p = stats.ttest_ind(arr1, arr2)
    # assign the t & p-value and assign significance mask to mask-column
    # ==> this_result_rowidx_mask is all False & one True, with the True being current
    #     contrast & SC Percentage
    this_result_rowidx_mask = (results_df[CONTRASTS_COL] == contrast) & (
        results_df[SC_PERCENTAGE_COL] == sc_percentage
    )
    results_df.loc[this_result_rowidx_mask, TTEST_T_COL] = this_t
    results_df.loc[this_result_rowidx_mask, TTEST_P_COL] = this_p
    if this_p < TTEST_MASK_THRESHOLD:
        results_df.loc[this_result_rowidx_mask, TTEST_MASK_COL] = True
    return results_df


def compute_and_assign_clustersize(results_df, contrast):
    """Compute size of all clusters of a given contrast and assign to results_df."""
    # prepare some variables
    ttest_mask_col_idx = results_df.columns.get_loc(TTEST_MASK_COL)
    ttest_tval_col_idx = results_df.columns.get_loc(TTEST_T_COL)
    tmass_col_idx = results_df.columns.get_loc(CLUSTER_TMASS_COL)
    this_tmass = 0.0  # tmass
    this_cluster_size = 0  # cluster size only used for checking if all ps were sig
    this_cluster_indices = []
    # loop over current contrast, update tmass, cluster size & indices if p was
    # significant
    for i in np.where(results_df[CONTRASTS_COL] == contrast)[0]:
        if results_df.iloc[i, ttest_mask_col_idx] == True:
            this_tmass += abs(results_df.iloc[i, ttest_tval_col_idx])  # tval
            this_cluster_size += 1  # cluster size
            this_cluster_indices.append(i)
            # handle the case of results_df ending with a sig. cluster
            if i == max(np.where(results_df[CONTRASTS_COL] == contrast)[0]):
                results_df.iloc[this_cluster_indices, tmass_col_idx] = this_tmass
        else:
            # if p was not significant, assign the previous cluster & reset our vars
            # => note this else also occurs when we keep having nonsig ts but for those
            #    nothing happens... coding it like this might make it a bit slower than
            #    a more sophisticated conditional logic here but the difference should
            #    be minimal so I just keep it as is
            results_df.iloc[this_cluster_indices, tmass_col_idx] = this_tmass
            this_tmass = 0.0
            this_cluster_size = 0
            this_cluster_indices = []
    # handle case of all ps being significant
    if this_cluster_size == len(np.where(results_df[CONTRASTS_COL] == contrast)[0]):
        results_df.iloc[this_cluster_indices, tmass_col_idx] = this_tmass
    return results_df


def extract_variable_array(df, stats_var, group_name, sc_percentage):
    """Extract an array of the variable we want to test with given ttest."""
    mask = (df[GROUP_COL] == group_name) & (df[SC_PERCENTAGE_COL] == sc_percentage)
    result = df.loc[mask, stats_var]
    return np.asarray(result)


# .........................  shuffle (permute) the true observed  ......................
def permute_true_observed_df(trueobs_df, cfg):
    """Shuffle groups of true observed and return permuted_df which is identical to
    trueobs_df except of GROUP_COL
    """
    # unpack
    bin_num = cfg["bin_num"]

    real_group_labels = list(trueobs_df.loc[0, GROUP_COL])
    permuted_group_labels = shuffle(real_group_labels)
    permuted_group_labels = np.repeat(permuted_group_labels, bin_num)
    permuted_df = trueobs_df.copy(deep=True)  # make sure to use copy here (not df=df)!!
    permuted_df.loc[:, GROUP_COL] = permuted_group_labels
    return permuted_df


# ................................  second-level test  .................................
def test_trueobs_clusters(
    trueobs_results_df, max_tmass, permutation_number, stats_threshold
):
    """Test the true observed cluster sizes against max cluster sizes under null."""
    # prepare stuff
    trueobs_results_df[CLUSTER_P_COL] = None
    trueobs_results_df[CLUSTER_MASK_COL] = False
    cluster_p_col_idx = trueobs_results_df.columns.get_loc(CLUSTER_P_COL)
    cluster_tmass_col_idx = trueobs_results_df.columns.get_loc(CLUSTER_TMASS_COL)
    cluster_mask_col_idx = trueobs_results_df.columns.get_loc(CLUSTER_MASK_COL)
    # loop over results, check each clustersize, assign final p value & mask
    for i in range(len(trueobs_results_df)):
        # tmass
        this_p = (
            sum(max_tmass >= trueobs_results_df.iloc[i, cluster_tmass_col_idx])
            / permutation_number
        )
        trueobs_results_df.iloc[i, cluster_p_col_idx] = this_p
        if this_p < stats_threshold:
            trueobs_results_df.iloc[i, cluster_mask_col_idx] = True
    return trueobs_results_df


# ...................................  plot results  ...................................
def plot_permutation_test_results(
    g_avg_dfs,
    g_std_dfs,
    trueobs_results_df,
    stats_var,
    folderinfo,
    cfg,
    plot_panel_instance,
):
    """Plot a Nx1 or N/2x2 figure of our contrasts' permutation test results."""

    # unpack
    group_names = folderinfo["group_names"]
    results_dir = folderinfo["results_dir"]
    contrasts = folderinfo["contrasts"]
    bin_num = cfg["bin_num"]
    group_color_dict = cfg["group_color_dict"]
    plot_SE = cfg["plot_SE"]
    feature = stats_var.split(" ")[-1]
    dont_show_plots = cfg["dont_show_plots"]
    legend_outside = cfg["legend_outside"]

    if len(contrasts) > 3:  # if we have 4 groups or more, N/2x2 subplot layout
        f, ax = plt.subplots(int(round(len(contrasts) / 2)), 2, layout="constrained")
        ax = ax.ravel()
    else:
        f, ax = plt.subplots(len(contrasts), 1, layout="constrained")
    x = np.linspace(0, 100, bin_num)
    for c, contrast in enumerate(contrasts):
        # prepare group strings and (importantly!) index of current groups from _NAMES
        groups = [group_name for group_name in contrast.split(" & ")]
        group_indices = [group_names.index(group_name) for group_name in groups]
        # plot observed g_avgs & g_stds
        for g, group_name in enumerate(groups):
            # group_idx is important. it correctly indexes dfs (!!) as well as colour!
            group_idx = group_indices[g]
            y_col = g_avg_dfs[group_idx].columns.get_loc(stats_var)
            y = g_avg_dfs[group_idx].iloc[:, y_col]
            if plot_SE:
                std = g_std_dfs[group_idx].iloc[:, y_col] / np.sqrt(
                    g_std_dfs[group_idx]["N"][0]
                )
            else:
                std = g_std_dfs[group_idx].iloc[:, y_col]
            this_color = group_color_dict[group_name]
            if type(ax) == np.ndarray:  # so we can do 2-group contrasts
                ax[c].plot(x, y, color=this_color, label=group_name, zorder=1)
                ax[c].fill_between(
                    x,
                    y - std,
                    y + std,
                    color=this_color,
                    alpha=STD_ALPHA,
                    lw=STD_LW,
                    zorder=1,
                )
            else:
                ax.plot(x, y, color=this_color, label=group_name, zorder=1)
                ax.fill_between(
                    x,
                    y - std,
                    y + std,
                    color=this_color,
                    alpha=STD_ALPHA,
                    lw=STD_LW,
                    zorder=1,
                )
        # convert to cm (if needed) before plotting clusters
        if type(ax) == np.ndarray:
            # legend adjustments
            if legend_outside is True:
                ax[c].legend(
                    fontsize=PERM_PLOT_LEGEND_SIZE,
                    loc="center left",
                    bbox_to_anchor=(1, 0.5),
                )
            elif legend_outside is False:
                ax[c].legend(fontsize=PERM_PLOT_LEGEND_SIZE)
            if check_mouse_conversion(feature, cfg):
                ytickconvert_mm_to_cm(feature, ax[c])
                ax[c].set_ylabel("")  # we use supylabel
        else:
            # legend adjustments
            if legend_outside is True:
                ax.legend(
                    fontsize=PERM_PLOT_LEGEND_SIZE + 4,
                    loc="center left",
                    bbox_to_anchor=(1, 0.5),
                )
            elif legend_outside is False:
                ax.legend(fontsize=PERM_PLOT_LEGEND_SIZE + 4)
        # plot significant clusters
        # => note that clusters is a list of list with idxs between 0 & bin_num-1
        clusters = extract_all_clusters(trueobs_results_df, contrast)
        if type(ax) == np.ndarray:
            ymin = ax[c].get_ylim()[0]
            ymax = ax[c].get_ylim()[1]
        else:
            ymin = ax.get_ylim()[0]
            ymax = ax.get_ylim()[1]
        for cluster in x[clusters]:  # index x with clusters == cluster has correct val
            x_coords = [cluster[0], cluster[1], cluster[1], cluster[0]]
            y_coords = [ymin, ymin, ymax, ymax]
            if type(ax) == np.ndarray:
                ax[c].fill(
                    x_coords,
                    y_coords,
                    color=BOX_COLOR,
                    alpha=BOX_ALPHA,
                    lw=STD_LW,
                    zorder=0,
                )
            else:
                ax.fill(
                    x_coords,
                    y_coords,
                    color=BOX_COLOR,
                    alpha=BOX_ALPHA,
                    lw=STD_LW,
                    zorder=0,
                )
    f.supxlabel("Percentage", fontsize=PERM_PLOT_SUPLABEL_SIZE)
    if check_mouse_conversion(feature, cfg):
        f.supylabel(feature + " (cm)", fontsize=PERM_PLOT_SUPLABEL_SIZE)
    else:
        f.supylabel(feature, fontsize=PERM_PLOT_SUPLABEL_SIZE)
    figure_file_string = stats_var + " - Cluster-extent Test"
    f.suptitle(figure_file_string, fontsize=PERM_PLOT_SUPLABEL_SIZE)
    save_figures(f, results_dir, figure_file_string)

    # add figure to plot panel figures list
    if dont_show_plots is False:  # -> show plot panel
        plot_panel_instance.figures.append(f)


def extract_all_clusters(trueobs_results_df, contrast):
    """Find indices of all (perm.) significant clusters"""
    contrast_mask = trueobs_results_df[CONTRASTS_COL] == contrast
    significance_mask = list(trueobs_results_df.loc[contrast_mask, CLUSTER_MASK_COL])
    all_clusters = []
    cluster = []
    for i, mask in enumerate(significance_mask):
        if mask == True:
            if len(cluster) == 0:
                cluster.append(i)
            if i == (len(significance_mask) - 1):
                cluster.append(i)
            else:
                if significance_mask[i + 1] == False:
                    cluster.append(i)
        if len(cluster) == 2:
            all_clusters.append(cluster)
            cluster = []
    return all_clusters


# %% .................  local functions #7 - 2-way RM/Mixed-ANOVA  .....................


# ...............................  sanity check . ......................................
def anova_design_sanity_check(stats_df, folderinfo, cfg):
    """Sanity check the anova_design input of the user based on stats_df's IDs"""

    # unpack
    anova_design = cfg["anova_design"]
    results_dir = folderinfo["results_dir"]

    # get IDs for each group to check if they are unique
    group_IDs = stats_df.groupby("Group")["ID"].unique()
    ID_list = []
    for this_groups_IDs in group_IDs:
        for ID in this_groups_IDs:
            ID_list.append(ID)
    ID_list = [str(IDs) for IDs in ID_list]
    unique_ID_list = list(set(ID_list))

    # Mixed ANOVA - no duplicate IDs across groups!
    if anova_design == "Mixed ANOVA":
        if len(ID_list) == len(unique_ID_list):  # check passed
            return True
        else:
            mixed_anova_error_message = (
                "\n*********\n! ERROR !\n*********\n"
                + "\nANOVA design seems wrong - skipping ANOVA!"
                + "\nMixed ANOVA requires unique IDs across groups & we found "
                + "duplicates!"
                + "\n\nIDs were:\n"
                + str(group_IDs)
            )
            print(mixed_anova_error_message)
            write_issues_to_textfile(mixed_anova_error_message, results_dir)
            return False
    # RM ANOVA - IDs in each group must be the same!
    elif anova_design == "RM ANOVA":
        if len(ID_list) != len(unique_ID_list):  # check passed
            # Bonus - inform user about which IDs were valid (have data)
            # ==> based on pingouin's approach of removing all IDs that do not have
            #     data in all conditions (see https://pingouin-stats.org/build/
            #     html/generated/pingouin.rm_anova.html under "Missing values...")
            valid_IDs = []
            group_number = len(group_IDs.index)
            for ID in unique_ID_list:
                ID_count = ID_list.count(ID)
                if ID_count == group_number:
                    valid_IDs.append(ID)
            rm_anova_info_message = (
                "\n********\n! INFO !\n********\n"
                + "\nFollowing IDs with valid data in all conditions after first-level "
                + "analyses will be included in RM ANOVA:\n\n"
                + str(valid_IDs)
            )
            print(rm_anova_info_message)
            write_issues_to_textfile(rm_anova_info_message, results_dir)
            return True
        else:
            rm_anova_error_message = (
                "\n*********\n! ERROR !\n*********\n"
                + "\nANOVA design seems wrong - skipping ANOVA!"
                + "\nRM ANOVA requires IDs to be present in all groups & we found "
                + "only unique IDs!"
                + "\n\nIDs were:\n"
                + str(group_IDs)
            )
            print(rm_anova_error_message)
            write_issues_to_textfile(rm_anova_error_message, results_dir)
            return False


# ...............................  main function  ......................................
def twoway_RMANOVA(
    stats_df, g_avg_dfs, g_std_dfs, stats_var, folderinfo, cfg, plot_panel_instance
):
    """Perform a two-way RM-ANOVA with the factors group (between or within) & SC
    percentage (within) on a given dependent variable
    """

    # unpack
    anova_design = cfg["anova_design"]

    # run the (fully) RM or Mixed ANOVA
    ANOVA_result = run_ANOVA(stats_df, stats_var, cfg)

    # check if sphericity is given to see if p vals have to be GG corrected
    # => even though this is done automatically for mixed anovas, you always get the GG
    #    col in results of rm anovas so we have to test it ourselves
    sphericity_flag = sphericity(
        stats_df, dv=stats_var, within=SC_PERCENTAGE_COL, subject=ID_COL
    )[0]
    if sphericity_flag:  # sphericity is given, no need to correct
        interaction_effect_pval = ANOVA_result["p-unc"][2]
    else:
        interaction_effect_pval = ANOVA_result["p-GG-corr"][2]
        # NU - understand why its nan sometimes. For now use uncorrected int pval
        if np.isnan(interaction_effect_pval):
            interaction_effect_pval = ANOVA_result["p-unc"][2]
    if interaction_effect_pval < 0.05:  # if interaction effect is sig, do multcomps
        multcomp_df = multcompare_SC_Percentages(stats_df, stats_var, folderinfo, cfg)
        save_stats_results_to_text(
            multcomp_df, stats_var, anova_design, folderinfo, cfg
        )
        plot_multcomp_results(
            g_avg_dfs,
            g_std_dfs,
            multcomp_df,
            stats_var,
            folderinfo,
            cfg,
            plot_panel_instance,
        )
    else:  # if interaction effect not sig, inform user that we didn't perform Tukey's!
        nonsig_multcomp_df = pd.DataFrame()
        save_stats_results_to_text(
            nonsig_multcomp_df, stats_var, "non-significant ANOVA", folderinfo, cfg
        )


# .............................  multiple comparison test  .............................
def multcompare_SC_Percentages(stats_df, stats_var, folderinfo, cfg):
    """Perform multiple comparison test if the ANOVA's interaction was significant.
    Do a separate multcomp test for each SC % bin."""

    # unpack
    group_names = folderinfo["group_names"]
    contrasts = folderinfo["contrasts"]
    bin_num = cfg["bin_num"]

    # prepare multcomp dataframe where we'll store results
    multcomp_df = pd.DataFrame(
        data=np.unique(stats_df[SC_PERCENTAGE_COL]),
        index=range(bin_num),
        columns=[SC_PERCENTAGE_COL],
    )
    multcomp_df = pd.concat(
        [multcomp_df, pd.DataFrame(data=None, index=range(bin_num), columns=contrasts)],
        axis=1,
    )
    # loop over SC Percentages & first prepare depvar_values of current SC % for testing
    for sc_perc in np.unique(stats_df[SC_PERCENTAGE_COL]):
        depvar_values = [[] for _ in range(len(group_names))]
        for g, group_name in enumerate(group_names):
            sc_perc_condition = stats_df[SC_PERCENTAGE_COL] == sc_perc
            group_condition = stats_df[GROUP_COL] == group_name
            mask = sc_perc_condition & group_condition
            depvar_values[g] = stats_df.loc[mask, stats_var].to_numpy()
        # perform the multcomps test and extract p values
        result = stats.tukey_hsd(*depvar_values)  # using * for group_num flexibility
        ps = result.pvalue
        # assign p values to multcomp results df
        # ==> see TukeyHSDResult class of scipy, according to their doc:
        #     "The element at index (i, j) is the p-value for the comparison between
        #     groups i and j." - so i & j matches contrasts as well as ps!
        sc_perc_row_idx = np.where(multcomp_df[SC_PERCENTAGE_COL] == sc_perc)[0][0]
        for i in range(len(group_names)):
            for j in range(i + 1, len(group_names)):
                contrast_col_idx = multcomp_df.columns.get_loc(
                    group_names[i] + " & " + group_names[j]
                )
                multcomp_df.iloc[sc_perc_row_idx, contrast_col_idx] = ps[i, j]
    return multcomp_df


# ..............................  main ANOVA  computation  .............................
def run_ANOVA(stats_df, stats_var, cfg):
    """Run the RM-ANOVA using pingouin"""

    # unpack
    anova_design = cfg["anova_design"]
    if anova_design == "Mixed ANOVA":
        result = stats_df.mixed_anova(
            dv=stats_var, between=GROUP_COL, within=SC_PERCENTAGE_COL, subject=ID_COL
        )
    elif anova_design == "RM ANOVA":
        result = stats_df.rm_anova(
            dv=stats_var, within=[GROUP_COL, SC_PERCENTAGE_COL], subject=ID_COL
        )
    return result


# ...................................  plot results  ...................................
def plot_multcomp_results(
    g_avg_dfs, g_std_dfs, multcomp_df, stats_var, folderinfo, cfg, plot_panel_instance
):
    """Plot an Nx1 figure of N contrasts' multiple comparison results."""

    # unpack
    group_names = folderinfo["group_names"]
    results_dir = folderinfo["results_dir"]
    contrasts = folderinfo["contrasts"]
    bin_num = cfg["bin_num"]
    group_color_dict = cfg["group_color_dict"]
    stats_threshold = cfg["stats_threshold"]
    plot_SE = cfg["plot_SE"]
    feature = stats_var.split(" ")[-1]
    dont_show_plots = cfg["dont_show_plots"]
    legend_outside = cfg["legend_outside"]

    f, ax = plt.subplots(len(contrasts), 1, layout="constrained")
    x = np.linspace(0, 100, bin_num)
    for c, contrast in enumerate(contrasts):
        # prepare group strings and (importantly!) index of current groups from _NAMES
        groups = [group_name for group_name in contrast.split(" & ")]
        group_indices = [group_names.index(group_name) for group_name in groups]
        # plot observed g_avgs & g_stds
        for g, group_name in enumerate(groups):
            # group_idx is important. it correctly indexes dfs (!!) as well as colour!
            group_idx = group_indices[g]
            y_col = g_avg_dfs[group_idx].columns.get_loc(stats_var)
            y = g_avg_dfs[group_idx].iloc[:, y_col]
            if plot_SE:
                std = g_std_dfs[group_idx].iloc[:, y_col] / np.sqrt(
                    g_std_dfs[group_idx]["N"][0]
                )
            else:
                std = g_std_dfs[group_idx].iloc[:, y_col]
            this_color = group_color_dict[group_name]
            if type(ax) == np.ndarray:  # so we can do a 2-way contrast
                ax[c].plot(x, y, color=this_color, label=group_name, zorder=1)
                ax[c].fill_between(
                    x,
                    y - std,
                    y + std,
                    color=this_color,
                    alpha=STD_ALPHA,
                    lw=STD_LW,
                    zorder=1,
                )
            else:
                ax.plot(x, y, color=this_color, label=group_name, zorder=1)
                ax.fill_between(
                    x,
                    y - std,
                    y + std,
                    color=this_color,
                    alpha=STD_ALPHA,
                    lw=STD_LW,
                    zorder=1,
                )
        # convert to cm (if needed) before plotting clusters
        if type(ax) == np.ndarray:
            # legend adjustments
            if legend_outside is True:
                ax[c].legend(
                    fontsize=PERM_PLOT_LEGEND_SIZE,
                    loc="center left",
                    bbox_to_anchor=(1, 0.5),
                )
            elif legend_outside is False:
                ax[c].legend(fontsize=PERM_PLOT_LEGEND_SIZE)
            if check_mouse_conversion(feature, cfg):
                ytickconvert_mm_to_cm(feature, ax[c])
                ax[c].set_ylabel("")  # we use supylabel
        else:
            # legend adjustments
            if legend_outside is True:
                ax.legend(
                    fontsize=PERM_PLOT_LEGEND_SIZE + 4,
                    loc="center left",
                    bbox_to_anchor=(1, 0.5),
                )
            elif legend_outside is False:
                ax.legend(fontsize=PERM_PLOT_LEGEND_SIZE + 4)
        # plot significant clusters
        clusters = extract_multcomp_significance_clusters(
            multcomp_df, contrast, stats_threshold
        )
        if type(ax) == np.ndarray:
            ymin = ax[c].get_ylim()[0]
            ymax = ax[c].get_ylim()[1]
        else:
            ymin = ax.get_ylim()[0]
            ymax = ax.get_ylim()[1]
        for cluster in x[clusters]:  # index x with clusters == cluster has correct val
            x_coords = [cluster[0], cluster[1], cluster[1], cluster[0]]
            y_coords = [ymin, ymin, ymax, ymax]
            if type(ax) == np.ndarray:
                ax[c].fill(
                    x_coords,
                    y_coords,
                    color=BOX_COLOR,
                    alpha=BOX_ALPHA,
                    lw=STD_LW,
                    zorder=0,
                )
            else:
                ax.fill(
                    x_coords,
                    y_coords,
                    color=BOX_COLOR,
                    alpha=BOX_ALPHA,
                    lw=STD_LW,
                    zorder=0,
                )
        if type(ax) == np.ndarray:
            # legend adjustments
            if legend_outside is True:
                ax[c].legend(
                    fontsize=PERM_PLOT_LEGEND_SIZE,
                    loc="center left",
                    bbox_to_anchor=(1, 0.5),
                )
            elif legend_outside is False:
                ax[c].legend(fontsize=PERM_PLOT_LEGEND_SIZE)
        else:
            # legend adjustments
            if legend_outside is True:
                ax.legend(
                    fontsize=PERM_PLOT_LEGEND_SIZE + 4,
                    loc="center left",
                    bbox_to_anchor=(1, 0.5),
                )
            elif legend_outside is False:
                ax.legend(fontsize=PERM_PLOT_LEGEND_SIZE + 4)
    f.supxlabel("Percentage", fontsize=PERM_PLOT_SUPLABEL_SIZE)
    if check_mouse_conversion(feature, cfg):
        f.supylabel(feature + " (cm)", fontsize=PERM_PLOT_SUPLABEL_SIZE)
    else:
        f.supylabel(feature, fontsize=PERM_PLOT_SUPLABEL_SIZE)
    figure_file_string = stats_var + " - Tukey's Multiple Comparison Test"
    f.suptitle(
        figure_file_string,
        fontsize=PERM_PLOT_SUPLABEL_SIZE,
    )
    save_figures(f, results_dir, figure_file_string)

    # add figure to plot panel figures list
    if dont_show_plots is False:  # -> show plot panel
        plot_panel_instance.figures.append(f)


def extract_multcomp_significance_clusters(multcomp_df, contrast, stats_threshold):
    """Extract clusters of significance after multiple comparison test"""
    # the df structure of this is different to permutation results df so we have to do
    # something slightly different here too
    significance_mask = multcomp_df[contrast] < stats_threshold
    all_clusters = []
    cluster = []
    for i, mask in enumerate(significance_mask):
        if mask == True:
            if len(cluster) == 0:
                cluster.append(i)
            if i == (len(significance_mask) - 1):
                cluster.append(i)
            else:
                if significance_mask[i + 1] == False:
                    cluster.append(i)
        if len(cluster) == 2:
            all_clusters.append(cluster)
            cluster = []
    return all_clusters


# %% ..........................  local functions #8 - plots  ...........................


def plot_results(g_avg_dfs, g_std_dfs, folderinfo, cfg, plot_panel_instance):
    """Plot results - main function (inner functions loop over groups)"""

    # unpack
    angles = cfg["angles"]
    tracking_software = cfg["tracking_software"]
    # prep
    plot_horizontal_coord = False

    # ........................1 - y coords over average SC..............................
    plot_joint_y_by_average_SC(
        g_avg_dfs, g_std_dfs, folderinfo, cfg, plot_panel_instance
    )
    plt.close("all")

    # .......................2 - x coords over average SC (optional)....................
    if tracking_software == "DLC":
        if cfg["analyse_average_x"] is True:
            plot_horizontal_coord = True
    elif tracking_software == "Simi":
        if cfg["analyse_average_y"] is True:
            plot_horizontal_coord = True
    if plot_horizontal_coord:
        plot_joint_x_by_average_SC(
            g_avg_dfs, g_std_dfs, folderinfo, cfg, plot_panel_instance
        )
        plt.close("all")

    # ........................3 - angles over average SC................................
    if angles["name"]:
        plot_angles_by_average_SC(
            g_avg_dfs, g_std_dfs, folderinfo, cfg, plot_panel_instance
        )
        plt.close("all")

    # .................4 - average x velocities over SC percentage......................
    plot_x_velocities_by_average_SC(
        g_avg_dfs, g_std_dfs, folderinfo, cfg, plot_panel_instance
    )
    plt.close("all")

    # ..............5 - average angular velocities over SC percentage...................
    if angles["name"]:
        plot_angular_velocities_by_average_SC(
            g_avg_dfs, g_std_dfs, folderinfo, cfg, plot_panel_instance
        )
        plt.close("all")

    # ........................optional - 6 - build plot panel..........................
    if cfg["dont_show_plots"] is True:
        pass  # going on without building the plot window
    elif cfg["dont_show_plots"] is False:  # -> show plot panel
        # Destroy loading screen and build plot panel with all figures
        plot_panel_instance.destroy_plot_panel_loading_screen()
        plot_panel_instance.build_plot_panel()


def plot_joint_y_by_average_SC(
    g_avg_dfs, g_std_dfs, folderinfo, cfg, plot_panel_instance
):
    """1 - Plot joints' y/Z as a function of average SC's percentage"""

    # unpack
    group_names = folderinfo["group_names"]
    results_dir = folderinfo["results_dir"]
    bin_num = cfg["bin_num"]
    joint_color_cycler = cfg["joint_color_cycler"]
    group_color_cycler = cfg["group_color_cycler"]
    which_leg = cfg["which_leg"]
    plot_SE = cfg["plot_SE"]
    tracking_software = cfg["tracking_software"]
    joints = cfg["joints"]
    dont_show_plots = cfg["dont_show_plots"]
    legend_outside = cfg["legend_outside"]

    # A - lines = joints & figures = groups
    for g, group_name in enumerate(group_names):
        f, ax = plt.subplots(1, 1)
        ax.set_prop_cycle(joint_color_cycler)
        x = np.linspace(0, 100, bin_num)
        for joint in joints:
            if tracking_software == "DLC":
                y_col = g_avg_dfs[g].columns.get_loc(joint + "y")
            elif tracking_software == "Simi":
                # check for bodyside-specificity
                feature = "Z"
                y_col = extract_feature_column(g_avg_dfs[g], joint, which_leg, feature)
            y = g_avg_dfs[g].iloc[:, y_col]  # average & stddata share colnames
            if plot_SE:
                std = g_std_dfs[g].iloc[:, y_col] / np.sqrt(g_std_dfs[g]["N"][0])
            else:
                std = g_std_dfs[g].iloc[:, y_col]
            ax.plot(x, y, label=joint)
            ax.fill_between(x, y - std, y + std, alpha=STD_ALPHA, lw=STD_LW)

        # legend adjustments
        if legend_outside is True:
            ax.legend(loc="center left", bbox_to_anchor=(1, 0.5))
        elif legend_outside is False:
            ax.legend()
        ax.set_xlabel("Percentage")
        if tracking_software == "DLC":
            ax.set_title(group_name + " - Joint Y over average step cycle")
            if check_mouse_conversion("y", cfg):
                ytickconvert_mm_to_cm("y", ax)
            else:
                ax.set_ylabel("y (pixel)")
            figure_file_string = " - Joint y-coord.s over average step cycle"
        elif tracking_software == "Simi":
            ax.set_title(
                group_name + " - " + which_leg + " Joint Z over average step cycle"
            )
            ax.set_ylabel("Z")
            figure_file_string = " - Joint Z-coord.s over average step cycle"
        save_figures(f, results_dir, group_name + figure_file_string)

        # add figure to plot panel figures list
        if dont_show_plots is False:  # -> show plot panel
            plot_panel_instance.figures.append(f)

    # B - lines = groups & figures = joints
    for j, joint in enumerate(joints):
        f, ax = plt.subplots(1, 1)
        ax.set_prop_cycle(group_color_cycler)
        x = np.linspace(0, 100, bin_num)
        for g, group_name in enumerate(group_names):
            if tracking_software == "DLC":
                y_col = g_avg_dfs[g].columns.get_loc(joint + "y")
            elif tracking_software == "Simi":
                # check for bodyside-specificity
                feature = "Z"
                y_col = extract_feature_column(g_avg_dfs[g], joint, which_leg, feature)
            y = g_avg_dfs[g].iloc[:, y_col]  # average & stddata share colnames
            if plot_SE:
                std = g_std_dfs[g].iloc[:, y_col] / np.sqrt(g_std_dfs[g]["N"][0])
            else:
                std = g_std_dfs[g].iloc[:, y_col]
            ax.plot(x, y, label=group_name)
            ax.fill_between(x, y - std, y + std, alpha=STD_ALPHA, lw=STD_LW)
        # legend adjustments
        if legend_outside is True:
            ax.legend(loc="center left", bbox_to_anchor=(1, 0.5))
        elif legend_outside is False:
            ax.legend()
        ax.set_xlabel("Percentage")
        if tracking_software == "DLC":
            ax.set_title(joint + "Y over average step cycle")
            if check_mouse_conversion("y", cfg):
                ytickconvert_mm_to_cm("y", ax)
            else:
                ax.set_ylabel("y (pixel)")
            figure_file_string = "- Y-coord.s over average step cycle"
        elif tracking_software == "Simi":
            # do title_leg thingy only for B plots, because here we have separate
            # figures for joints / angles (in A plots just throw leg into title)
            if joint + "Z" in g_avg_dfs[g].columns:
                title_leg = ""
            else:
                title_leg = which_leg
            ax.set_title(title_leg + " " + joint + " Z over average step cycle")
            ax.set_ylabel("Z")
            figure_file_string = "- Z-coord.s over average step cycle"
        save_figures(f, results_dir, joint + figure_file_string)

        # add figure to plot panel figures list
        if dont_show_plots is False:  # -> show plot panel
            plot_panel_instance.figures.append(f)


def plot_joint_x_by_average_SC(
    g_avg_dfs, g_std_dfs, folderinfo, cfg, plot_panel_instance
):
    """2 - Plot joints' x/Y as a function of average SC's percentage"""

    # unpack
    group_names = folderinfo["group_names"]
    results_dir = folderinfo["results_dir"]
    bin_num = cfg["bin_num"]
    joint_color_cycler = cfg["joint_color_cycler"]
    group_color_cycler = cfg["group_color_cycler"]
    which_leg = cfg["which_leg"]
    plot_SE = cfg["plot_SE"]
    tracking_software = cfg["tracking_software"]
    joints = cfg["joints"]
    dont_show_plots = cfg["dont_show_plots"]
    legend_outside = cfg["legend_outside"]

    # A - lines = joints & figures = groups
    for g, group_name in enumerate(group_names):
        f, ax = plt.subplots(1, 1)
        ax.set_prop_cycle(joint_color_cycler)
        x = np.linspace(0, 100, bin_num)
        for joint in joints:
            if tracking_software == "DLC":
                y_col = g_avg_dfs[g].columns.get_loc(joint + "x")
            elif tracking_software == "Simi":
                # check for bodyside-specificity
                feature = "Y"
                y_col = extract_feature_column(g_avg_dfs[g], joint, which_leg, feature)
            y = g_avg_dfs[g].iloc[:, y_col]  # average & stddata share colnames
            if plot_SE:
                std = g_std_dfs[g].iloc[:, y_col] / np.sqrt(g_std_dfs[g]["N"][0])
            else:
                std = g_std_dfs[g].iloc[:, y_col]
            ax.plot(x, y, label=joint)
            ax.fill_between(x, y - std, y + std, alpha=STD_ALPHA, lw=STD_LW)

        # legend adjustments
        if legend_outside is True:
            ax.legend(loc="center left", bbox_to_anchor=(1, 0.5))
        elif legend_outside is False:
            ax.legend()
        ax.set_xlabel("Percentage")
        if tracking_software == "DLC":
            ax.set_title(group_name + " - Joint X over average step cycle")
            if check_mouse_conversion("x", cfg):
                ytickconvert_mm_to_cm("x", ax)
            else:
                ax.set_ylabel("x (pixel)")
            figure_file_string = " - Joint x-coord.s over average step cycle"
        elif tracking_software == "Simi":
            ax.set_title(
                group_name + " - " + which_leg + " Joint Y over average step cycle"
            )
            ax.set_ylabel("Y")
            figure_file_string = " - Joint Y-coord.s over average step cycle"
        save_figures(f, results_dir, group_name + figure_file_string)

        # add figure to plot panel figures list
        if dont_show_plots is False:  # -> show plot panel
            plot_panel_instance.figures.append(f)

    # B - lines = groups & figures = joints
    for j, joint in enumerate(joints):
        f, ax = plt.subplots(1, 1)
        ax.set_prop_cycle(group_color_cycler)
        x = np.linspace(0, 100, bin_num)
        for g, group_name in enumerate(group_names):
            if tracking_software == "DLC":
                y_col = g_avg_dfs[g].columns.get_loc(joint + "x")
            elif tracking_software == "Simi":
                # check for bodyside-specificity
                feature = "Y"
                y_col = extract_feature_column(g_avg_dfs[g], joint, which_leg, feature)
            y = g_avg_dfs[g].iloc[:, y_col]  # average & stddata share colnames
            if plot_SE:
                std = g_std_dfs[g].iloc[:, y_col] / np.sqrt(g_std_dfs[g]["N"][0])
            else:
                std = g_std_dfs[g].iloc[:, y_col]
            ax.plot(x, y, label=group_name)
            ax.fill_between(x, y - std, y + std, alpha=STD_ALPHA, lw=STD_LW)
        # legend adjustments
        if legend_outside is True:
            ax.legend(loc="center left", bbox_to_anchor=(1, 0.5))
        elif legend_outside is False:
            ax.legend()
        ax.set_xlabel("Percentage")
        if tracking_software == "DLC":
            ax.set_title(joint + "X over average step cycle")
            if check_mouse_conversion("x", cfg):
                ytickconvert_mm_to_cm("x", ax)
            else:
                ax.set_ylabel("x (pixel)")
            figure_file_string = "- X-coord.s over average step cycle"
        elif tracking_software == "Simi":
            # do title_leg thingy only for B plots, because here we have separate
            # figures for joints / angles (in A plots just throw leg into title)
            if joint + "Y" in g_avg_dfs[g].columns:
                title_leg = ""
            else:
                title_leg = which_leg
            ax.set_title(title_leg + " " + joint + " Y over average step cycle")
            ax.set_ylabel("Y")
            figure_file_string = "- Y-coord.s over average step cycle"
        save_figures(f, results_dir, joint + figure_file_string)

        # add figure to plot panel figures list
        if dont_show_plots is False:  # -> show plot panel
            plot_panel_instance.figures.append(f)


def plot_angles_by_average_SC(
    g_avg_dfs, g_std_dfs, folderinfo, cfg, plot_panel_instance
):
    """2 - Plot Angles as a function of average SC's percentage"""

    # unpack
    group_names = folderinfo["group_names"]
    results_dir = folderinfo["results_dir"]
    bin_num = cfg["bin_num"]
    angle_color_cycler = cfg["angle_color_cycler"]
    group_color_cycler = cfg["group_color_cycler"]
    which_leg = cfg["which_leg"]
    plot_SE = cfg["plot_SE"]
    tracking_software = cfg["tracking_software"]
    angles = cfg["angles"]
    dont_show_plots = cfg["dont_show_plots"]
    legend_outside = cfg["legend_outside"]

    # A - lines = angles & figures = groups
    for g, group_name in enumerate(group_names):
        f, ax = plt.subplots(1, 1)
        ax.set_prop_cycle(angle_color_cycler)
        x = np.linspace(0, 100, bin_num)
        for angle in angles["name"]:
            if tracking_software == "DLC":
                y_col = g_avg_dfs[g].columns.get_loc(angle + "Angle")
            elif tracking_software == "Simi":
                # check for bodyside-specificity
                feature = "Angle"
                y_col = extract_feature_column(g_avg_dfs[g], angle, which_leg, feature)
            y = g_avg_dfs[g].iloc[:, y_col]  # sharing colnames
            if plot_SE:
                std = g_std_dfs[g].iloc[:, y_col] / np.sqrt(g_std_dfs[g]["N"][0])
            else:
                std = g_std_dfs[g].iloc[:, y_col]
            ax.plot(x, y, label=angle)
            ax.fill_between(x, y - std, y + std, alpha=STD_ALPHA, lw=STD_LW)
        # legend adjustments
        if legend_outside is True:
            ax.legend(loc="center left", bbox_to_anchor=(1, 0.5))
        elif legend_outside is False:
            ax.legend()
        ax.set_xlabel("Percentage")
        ax.set_ylabel("Angle (degrees)")
        if tracking_software == "DLC":
            ax.set_title(group_name + " - Joint angles over average step cycle")
        elif tracking_software == "Simi":
            ax.set_title(
                group_name + " - " + which_leg + " joint angles over average step cycle"
            )
        figure_file_string = " - Joint angles over average step cycle"
        save_figures(f, results_dir, group_name + figure_file_string)

        # add figure to plot panel figures list
        if dont_show_plots is False:  # -> show plot panel
            plot_panel_instance.figures.append(f)

    # B - lines = groups & figures = angles
    for a, angle in enumerate(angles["name"]):
        f, ax = plt.subplots(1, 1)
        ax.set_prop_cycle(group_color_cycler)
        x = np.linspace(0, 100, bin_num)
        for g, group_name in enumerate(group_names):
            if tracking_software == "DLC":
                y_col = g_avg_dfs[g].columns.get_loc(angle + "Angle")
            elif tracking_software == "Simi":
                # check for bodyside-specificity
                feature = "Angle"
                y_col = extract_feature_column(g_avg_dfs[g], angle, which_leg, feature)
            y = g_avg_dfs[g].iloc[:, y_col]  # average & stddata share colnames
            if plot_SE:
                std = g_std_dfs[g].iloc[:, y_col] / np.sqrt(g_std_dfs[g]["N"][0])
            else:
                std = g_std_dfs[g].iloc[:, y_col]
            ax.plot(x, y, label=group_name)
            ax.fill_between(x, y - std, y + std, alpha=STD_ALPHA, lw=STD_LW)
        # legend adjustments
        if legend_outside is True:
            ax.legend(loc="center left", bbox_to_anchor=(1, 0.5))
        elif legend_outside is False:
            ax.legend()
        ax.set_xlabel("Percentage")
        ax.set_ylabel("Angle (degrees)")
        if tracking_software == "DLC":
            ax.set_title(angle + "angle over average step cycle")
        elif tracking_software == "Simi":
            if angle + "Angle" in g_avg_dfs[g].columns:
                title_leg = ""
            else:
                title_leg = which_leg
            ax.set_title(title_leg + " " + angle + " angle over average step cycle")
        figure_file_string = " - Angle over average step cycle"
        save_figures(f, results_dir, angle + figure_file_string)

        # add figure to plot panel figures list
        if dont_show_plots is False:  # -> show plot panel
            plot_panel_instance.figures.append(f)


def plot_x_velocities_by_average_SC(
    g_avg_dfs, g_std_dfs, folderinfo, cfg, plot_panel_instance
):
    """3 - Plot x velocities as a function of average SC's percentage"""

    # unpack
    group_names = folderinfo["group_names"]
    results_dir = folderinfo["results_dir"]
    bin_num = cfg["bin_num"]
    joint_color_cycler = cfg["joint_color_cycler"]
    group_color_cycler = cfg["group_color_cycler"]
    which_leg = cfg["which_leg"]
    sampling_rate = cfg["sampling_rate"]
    plot_SE = cfg["plot_SE"]
    tracking_software = cfg["tracking_software"]
    joints = cfg["joints"]
    dont_show_plots = cfg["dont_show_plots"]
    legend_outside = cfg["legend_outside"]

    # A - lines = joints & figures = groups
    for g, group_name in enumerate(group_names):
        f, ax = plt.subplots(1, 1)
        ax.set_prop_cycle(joint_color_cycler)
        x = np.linspace(0, 100, bin_num)
        for joint in joints:
            if tracking_software == "DLC":
                y_col = g_avg_dfs[g].columns.get_loc(joint + "Velocity")
            elif tracking_software == "Simi":
                # check for bodyside-specificity
                feature = "Velocity"
                y_col = extract_feature_column(g_avg_dfs[g], joint, which_leg, feature)
            y = g_avg_dfs[g].iloc[:, y_col]
            if plot_SE:
                std = g_std_dfs[g].iloc[:, y_col] / np.sqrt(g_std_dfs[g]["N"][0])
            else:
                std = g_std_dfs[g].iloc[:, y_col]
            ax.plot(x, y, label=joint)
            ax.fill_between(x, y - std, y + std, alpha=0.2)
        # legend adjustments
        if legend_outside is True:
            ax.legend(loc="center left", bbox_to_anchor=(1, 0.5))
        elif legend_outside is False:
            ax.legend()
        ax.set_xlabel("Percentage")
        if tracking_software == "DLC":
            if check_mouse_conversion("x", cfg):
                ytickconvert_mm_to_cm("x", ax)
                ax.set_ylabel(  # overwriting convert-funcs ylabel
                    "Velocity (x in cm / "
                    + str(int((1 / sampling_rate) * 1000))
                    + " ms)"
                )
            else:
                ax.set_ylabel(
                    "Velocity (x in pixels / "
                    + str(int((1 / sampling_rate) * 1000))
                    + " ms)"
                )
            ax.set_title(group_name + " - Joint velocities over average step cycle")
        elif tracking_software == "Simi":
            ax.set_ylabel(
                "Velocity (Y in (your_units) / "
                + str(int((1 / sampling_rate) * 1000))
                + " ms)"
            )
            ax.set_title(
                group_name
                + " - "
                + which_leg
                + " joint velocities over average step cycle"
            )
        figure_file_string = " - Joint velocities over average step cycle"
        save_figures(f, results_dir, group_name + figure_file_string)

        # add figure to plot panel figures list
        if dont_show_plots is False:  # -> show plot panel
            plot_panel_instance.figures.append(f)

    # B - lines = groups & figures = joints
    for j, joint in enumerate(joints):
        f, ax = plt.subplots(1, 1)
        ax.set_prop_cycle(group_color_cycler)
        x = np.linspace(0, 100, bin_num)
        for g, group_name in enumerate(group_names):
            if tracking_software == "DLC":
                y_col = g_avg_dfs[g].columns.get_loc(joint + "Velocity")
            elif tracking_software == "Simi":
                # check for bodyside-specificity
                feature = "Velocity"
                y_col = extract_feature_column(g_avg_dfs[g], joint, which_leg, feature)
            y = g_avg_dfs[g].iloc[:, y_col]  # average & stddata share colnames
            if plot_SE:
                std = g_std_dfs[g].iloc[:, y_col] / np.sqrt(g_std_dfs[g]["N"][0])
            else:
                std = g_std_dfs[g].iloc[:, y_col]
            ax.plot(x, y, label=group_name)
            ax.fill_between(x, y - std, y + std, alpha=STD_ALPHA, lw=STD_LW)
        # legend adjustments
        if legend_outside is True:
            ax.legend(loc="center left", bbox_to_anchor=(1, 0.5))
        elif legend_outside is False:
            ax.legend()
        ax.set_xlabel("Percentage")
        if tracking_software == "DLC":
            if check_mouse_conversion("x", cfg):
                ytickconvert_mm_to_cm("x", ax)
                ax.set_ylabel(
                    "Velocity (x in cm / "
                    + str(int((1 / sampling_rate) * 1000))
                    + " ms)"
                )
            else:
                ax.set_ylabel(
                    "Velocity (x in pixels / "
                    + str(int((1 / sampling_rate) * 1000))
                    + " ms)"
                )
            ax.set_title(joint + "velocities over average step cycle")
        elif tracking_software == "Simi":
            ax.set_ylabel(
                "Velocity (Y in (your_units) / "
                + str(int((1 / sampling_rate) * 1000))
                + " ms)"
            )
            if joint + "Velocity" in g_avg_dfs[g].columns:
                title_leg = ""
            else:
                title_leg = which_leg
            ax.set_title(
                title_leg + " " + joint + " velocities over average step cycle"
            )
        figure_file_string = "- Velocities over average step cycle"
        save_figures(f, results_dir, joint + figure_file_string)

        # add figure to plot panel figures list
        if dont_show_plots is False:  # -> show plot panel
            plot_panel_instance.figures.append(f)


def plot_angular_velocities_by_average_SC(
    g_avg_dfs, g_std_dfs, folderinfo, cfg, plot_panel_instance
):
    """4 - Plot angular velocities as a function of average SC's percentage"""
    # unpack
    group_names = folderinfo["group_names"]
    results_dir = folderinfo["results_dir"]
    bin_num = cfg["bin_num"]
    angle_color_cycler = cfg["angle_color_cycler"]
    group_color_cycler = cfg["group_color_cycler"]
    which_leg = cfg["which_leg"]
    sampling_rate = cfg["sampling_rate"]
    plot_SE = cfg["plot_SE"]
    tracking_software = cfg["tracking_software"]
    angles = cfg["angles"]
    dont_show_plots = cfg["dont_show_plots"]
    legend_outside = cfg["legend_outside"]

    # A - lines = joints & figures = groups
    for g, group_name in enumerate(group_names):
        f, ax = plt.subplots(1, 1)
        ax.set_prop_cycle(angle_color_cycler)
        x = np.linspace(0, 100, bin_num)
        for angle in angles["name"]:
            if tracking_software == "DLC":
                y_col = g_avg_dfs[g].columns.get_loc(angle + "Angle Velocity")
            elif tracking_software == "Simi":
                # check for bodyside-specificity
                feature = "Angle Velocity"
                y_col = extract_feature_column(g_avg_dfs[g], angle, which_leg, feature)
            y = g_avg_dfs[g].iloc[:, y_col]
            if plot_SE:
                std = g_std_dfs[g].iloc[:, y_col] / np.sqrt(g_std_dfs[g]["N"][0])
            else:
                std = g_std_dfs[g].iloc[:, y_col]
            ax.plot(x, y, label=angle)
            ax.fill_between(x, y - std, y + std, alpha=0.2)
        # legend adjustments
        if legend_outside is True:
            ax.legend(loc="center left", bbox_to_anchor=(1, 0.5))
        elif legend_outside is False:
            ax.legend()
        ax.set_xlabel("Percentage")
        ax.set_ylabel(
            "Velocity (degree / " + str(int((1 / sampling_rate) * 1000)) + " ms)"
        )
        if tracking_software == "DLC":
            ax.set_title(group_name + " - Angular velocities over average step cycle")
        elif tracking_software == "Simi":
            ax.set_title(
                group_name
                + " - "
                + which_leg
                + " angular velocities over average step cycle"
            )
        figure_file_string = " - Angular velocities over average step cycle"
        save_figures(f, results_dir, group_name + figure_file_string)

        # add figure to plot panel figures list
        if dont_show_plots is False:  # -> show plot panel
            plot_panel_instance.figures.append(f)

    # B - lines = groups & figures = joints
    for a, angle in enumerate(angles["name"]):
        f, ax = plt.subplots(1, 1)
        ax.set_prop_cycle(group_color_cycler)
        x = np.linspace(0, 100, bin_num)
        for g, group_name in enumerate(group_names):
            if tracking_software == "DLC":
                y_col = g_avg_dfs[g].columns.get_loc(angle + "Angle Velocity")
            elif tracking_software == "Simi":
                # check for bodyside-specificity
                feature = "Angle Velocity"
                y_col = extract_feature_column(g_avg_dfs[g], angle, which_leg, feature)
            y = g_avg_dfs[g].iloc[:, y_col]  # average & stddata share colnames
            if plot_SE:
                std = g_std_dfs[g].iloc[:, y_col] / np.sqrt(g_std_dfs[g]["N"][0])
            else:
                std = g_std_dfs[g].iloc[:, y_col]
            ax.plot(x, y, label=group_name)
            ax.fill_between(x, y - std, y + std, alpha=STD_ALPHA, lw=STD_LW)
        # legend adjustments
        if legend_outside is True:
            ax.legend(loc="center left", bbox_to_anchor=(1, 0.5))
        elif legend_outside is False:
            ax.legend()
        ax.set_xlabel("Percentage")
        ax.set_ylabel(
            "Velocity (degree / " + str(int((1 / sampling_rate) * 1000)) + " ms)"
        )
        if tracking_software == "DLC":
            ax.set_title(angle + "- Angular velocities over average step cycle")
        elif tracking_software == "Simi":
            if angle + "Angle" in g_avg_dfs[g].columns:
                title_leg = ""
            else:
                title_leg = which_leg
            ax.set_title(
                title_leg
                + " "
                + angle
                + " - Angular velocities over average step cycle"
            )
        figure_file_string = " - Angular Velocities over average step cycle"
        save_figures(f, results_dir, angle + figure_file_string)

        # add figure to plot panel figures list
        if dont_show_plots is False:  # -> show plot panel
            plot_panel_instance.figures.append(f)


def save_figures(figure, results_dir, figure_file_string):
    """Save figures as pngs to results_dir and as svgs to separate subfolders"""
    # pngs to results_dir
    figure.savefig(
        os.path.join(results_dir, figure_file_string + ".png"),
        bbox_inches="tight",
    )
    # svgs to subfolders
    svg_dir = os.path.join(results_dir, "SVG Figures")
    if not os.path.exists(svg_dir):
        os.makedirs(svg_dir)
    figure.savefig(
        os.path.join(svg_dir, figure_file_string + ".svg"), bbox_inches="tight"
    )


def ytickconvert_mm_to_cm(feature, axis):
    """Convert axis y-ticks from mm (of data) to cm"""
    y_ticks = axis.get_yticks()
    y_ticklabels = []
    for t in y_ticks:
        y_ticklabels.append(str(round(t / 10, 2)))
    axis.set_yticks(y_ticks, labels=y_ticklabels)
    axis.set_ylabel(feature + " (cm)")


def extract_feature_column(df, joint, legname, feature):
    """Extract the column of a given joint (or angle) x legname (or not) x feature combo

    Note
    ----
    Only use this when using .iloc, not .loc!
    ==> the return statement gives the column-index to be used with .iloc!
    """
    if joint + feature in df.columns:
        string = joint + feature
    else:
        string = transform_joint_and_leg_to_colname(joint, legname, feature)
    return df.columns.get_loc(string)


# %% .............  local functions #0 & #9 - print start & finish  ....................


def print_start(folderinfo, cfg):
    """Print some info about starting this analysis"""

    # header
    start_string = (
        "\n******************\n AutoGaitA_Group \n******************"
        + "\n\nContrasting Groups:"
    )
    # groups
    for group_name in folderinfo["group_names"]:
        start_string += "\n" + group_name
    # pca
    start_string += "\n\n\n*****\n PCA \n*****"
    if cfg["PCA_variables"]:
        start_string += "\n\nFeatures:"
        for PCA_var in cfg["PCA_variables"]:
            start_string += "\n" + PCA_var
        start_string += (
            "\n\nConfiguration:\n" + str(cfg["number_of_PCs"]) + " principal components"
        )
    else:
        start_string += "\n\nNo PCA wanted!"
    # stats
    start_string += "\n\n\n*************\n Statistics \n*************"
    if cfg["stats_variables"]:
        start_string += "\n\nFeatures:"
        for stats_var in cfg["stats_variables"]:
            start_string += "\n" + stats_var
        start_string += "\n\nConfiguration:"
        if cfg["do_anova"]:
            start_string += "\n" + cfg["anova_design"]
        else:
            start_string += "\nNo Anova"
        start_string += (
            "\nCluster-extent permutation test with "
            + str(cfg["permutation_number"])
            + " permutations"
        )
        start_string += (
            "\nAlpha Level of " + str(cfg["stats_threshold"] * 100) + "%\n\n"
        )
    else:
        start_string += "\n\nNo stats wanted!\n\n"

    # done - print
    print(start_string)


def print_finish(folderinfo):
    """Inform the user about being done."""
    print("\n***************************************************************")
    print("*      GAITA FINISHED - YOUR RESULTS WERE STORED HERE:        *")
    print(folderinfo["results_dir"])
    print("***************************************************************")


# %% ................  local functions #10 - misc. helper functions  ...................


def transform_joint_and_leg_to_colname(joint, legname, feature):
    """For Human Data: Transform a joint and leg name to Simi-column name"""
    return joint + ", " + legname + " " + feature


def write_issues_to_textfile(message, results_dir):
    """Write issues to a textfile"""
    issues_textfile = os.path.join(results_dir, ISSUES_TXT_FILENAME)
    with open(issues_textfile, "a") as f:
        f.write(message)


def check_mouse_conversion(feature, cfg):
    """For mouse plots: check if we have to convert mm to cm (for plotting x or y)"""
    if "convert_to_mm" not in cfg.keys():
        return False
    else:
        if cfg["convert_to_mm"] is False:
            return False
        else:
            if (
                feature.endswith(" x")
                | feature.endswith(" y")
                | (feature in ["x", "y"])
            ):
                return True


def save_stats_results_to_text(results_df, stats_var, which_test, folderinfo, cfg):
    """Save the numerical results of our cluster extent or ANOVA results to a text file
    Note
    ----
    which_test can either be:
        "RM ANOVA", "Mixed ANOVA", "non-significant ANOVA",
        or "Cluster-extent permutation test"
        If in the future you want to have some other test be mindful about the:
            if "ANOVA" in which_test lines!
    """
    # unpack
    contrasts = folderinfo["contrasts"]
    results_dir = folderinfo["results_dir"]
    bin_num = cfg["bin_num"]
    stats_threshold = cfg["stats_threshold"]

    # initial message
    message = (
        "\n\n**************************************************************************"
        + "\n\n*****  Results of "
        + which_test
        + " for "
        + stats_var
        + "  ****"
    )

    # contrast specific info
    for contrast in contrasts:
        # extract significant clusters
        # => works slightly different based on which statistical test we used, but in
        #    both cases it returns a list of lists of indices with range(bin_num)
        #   - which is why we can use rounded_sc_percentages as we do below
        if "ANOVA" in which_test:
            if "non-significant" in which_test:  # interaction nonsig - no multcomps
                clusters = []
            else:
                # interaction significant (which_test is either RM ANOVA or Mixed ANOVA)
                clusters = extract_multcomp_significance_clusters(
                    results_df, contrast, stats_threshold
                )
        else:
            clusters = extract_all_clusters(results_df, contrast)
        # write message
        if len(clusters) == 0:  # no sig clusters were found!
            if "non-significant" in which_test:
                message = (
                    message
                    + "\n\nContrast: "
                    + contrast
                    + " - interaction effect "
                    + "not significant. Tukey's test was not performed!"
                )
            else:
                message = (
                    message
                    + "\n\nContrast: "
                    + contrast
                    + " - No significant clusters"
                    + "!"
                )
        else:
            rounded_sc_percentages = np.linspace(0, 100, bin_num).round(2)
            message = (
                message
                + "\n\nContrast: "
                + contrast
                + " - Significant clusters found at (in SC Percentage):"
            )
            for cluster in clusters:
                message = (
                    message
                    + "\n\n"
                    + str(rounded_sc_percentages[cluster[0]])
                    + "-"
                    + str(rounded_sc_percentages[cluster[1]])
                    + "%"
                    + ", p values:\n\n"
                )
                # add pvals to message
                # => handle ANOVA and perm test differently since results_df different
                if "ANOVA" in which_test:
                    # important to include cluster-end bin here thus cluster[1]+1!
                    for i in range(cluster[0], cluster[1] + 1):
                        message = (
                            message
                            + "\n"
                            + str(rounded_sc_percentages[i])
                            + "% - p = "
                            + str(round(results_df.loc[i, contrast], 4))
                        )
                else:
                    # extract subset df of only this contrast because cluster variables
                    # idxs values correspond to 0:bin_num and results_df of perm test
                    # (this is different for anovas) has 0:bin_num*contrast_number!
                    this_contrast_results_df = results_df[
                        results_df[CONTRASTS_COL] == contrast
                    ]
                    cluster_p_colidx = this_contrast_results_df.columns.get_loc(
                        CLUSTER_P_COL
                    )
                    # also note that we only use cluster[0] to retrieve cluster's pval
                    # since the pval is constant across the whole cluster always
                    message = (
                        message
                        + "Cluster p value = "
                        + str(
                            this_contrast_results_df.iloc[cluster[0], cluster_p_colidx]
                        )
                    )

    # message end
    message = (
        message
        + "\n\n***********************************************************************"
        + "***\n\n"
    )

    # print & save
    print(message)
    stats_textfile = os.path.join(results_dir, STATS_TXT_FILENAME)
    with open(stats_textfile, "a") as f:
        f.write(message)


class PlotPanel:
    def __init__(self):
        self.figures = []
        self.current_fig_index = 0

    # .........................  loading screen  ................................
    def build_plot_panel_loading_screen(self):
        """Builds a loading screen that is shown while plots are generated"""
        # Build window
        self.loading_screen = ctk.CTkToplevel()
        self.loading_screen.title("Loading...")
        self.loading_screen.geometry("300x300")
        self.loading_label_strings = [
            "Plots are generated, please wait.",
            "Plots are generated, please wait..",
            "Plots are generated, please wait...",
        ]
        self.loading_label = ctk.CTkLabel(
            self.loading_screen, text=self.loading_label_strings[0]
        )
        self.loading_label.pack(pady=130, padx=40, anchor="w")

        # Animate the text
        self.animate(counter=1)

    # Cycle through loading labels to animate the loading screen
    def animate(self, counter):
        self.loading_label.configure(text=self.loading_label_strings[counter])
        self.loading_screen.after(
            500, self.animate, (counter + 1) % len(self.loading_label_strings)
        )

    def destroy_plot_panel_loading_screen(self):
        self.loading_screen.destroy()

    # .........................  plot panel   ................................
    def build_plot_panel(self):
        """Creates the window/"panel" in which the plots are shown"""
        # Set up of the plotpanel
        ctk.set_appearance_mode("dark")  # Modes: system (default), light, dark
        ctk.set_default_color_theme("green")  # Themes: blue , dark-blue, green
        self.plotwindow = ctk.CTkToplevel()
        self.plotwindow.title(
            f"AutoGaitA Figure {self.current_fig_index+1}/{len(self.figures)}"
        )

        # Set size to 50% of screen
        screen_width = self.plotwindow.winfo_screenwidth()
        window_width = int(screen_width * 0.5)
        # 0.75 to gain a ration of 1.333 (that of matplotlib figures) and 1.05 for toolbar + buttons
        window_height = window_width * 0.75 * 1.05
        self.plotwindow.geometry(f"{window_width}x{window_height}")

        # Adjust figures for the plot panel
        for fig in self.figures:
            # dpi adjusted to increase visibilty/readability
            fig.set_dpi(100)
            # constrained layout to adjust margins within the figure
            # => note: in case there are a lot of steps in one run (-> the legend is
            #          super long) the figure won't be displayed properly.
            fig.set_constrained_layout(True)

        # Initialize the plot panel with the first figure
        self.plot_panel = FigureCanvasTkAgg(
            self.figures[self.current_fig_index], master=self.plotwindow
        )  # index used for buttons
        self.plot_panel.get_tk_widget().grid(
            row=0, column=0, padx=10, pady=10, sticky="nsew"
        )

        # Create toolbar frame and place it in the middle row
        self.toolbar_frame = tk.Frame(self.plotwindow)
        self.toolbar_frame.grid(row=1, column=0, sticky="ew")

        self.toolbar = NavigationToolbar2Tk(self.plot_panel, self.toolbar_frame)
        self.toolbar.update()

        # Create navigation buttons frame
        self.button_frame = tk.Frame(self.plotwindow)
        self.button_frame.grid(row=2, column=0, sticky="ew")

        self.prev_button = ctk.CTkButton(
            self.button_frame,
            text="<< Previous",
            fg_color=FG_COLOR,
            hover_color=HOVER_COLOR,
            command=self.show_previous,
        )
        self.next_button = ctk.CTkButton(
            self.button_frame,
            text="Next >>",
            fg_color=FG_COLOR,
            hover_color=HOVER_COLOR,
            command=self.show_next,
        )
        self.prev_button.grid(row=0, column=0, sticky="ew")
        self.next_button.grid(row=0, column=1, sticky="ew")

        self.button_frame.grid_columnconfigure(0, weight=1)
        self.button_frame.grid_columnconfigure(1, weight=1)

        # Configure grid layout
        self.plotwindow.grid_rowconfigure(0, weight=1)
        self.plotwindow.grid_rowconfigure(1, weight=0)
        self.plotwindow.grid_rowconfigure(2, weight=0)
        self.plotwindow.grid_columnconfigure(0, weight=1)

    def show_previous(self):
        if self.current_fig_index > 0:
            self.current_fig_index -= 1
            self.update_plot_and_toolbar()

    def show_next(self):
        if self.current_fig_index < len(self.figures) - 1:
            self.current_fig_index += 1
            self.update_plot_and_toolbar()

    def update_plot_and_toolbar(self):
        # Clear the current plot panel
        self.plot_panel.get_tk_widget().grid_forget()

        # Update the plot panel with the new figure
        self.plot_panel = FigureCanvasTkAgg(
            self.figures[self.current_fig_index], master=self.plotwindow
        )
        self.plot_panel.get_tk_widget().grid(row=0, column=0, sticky="nsew")
        self.plot_panel.draw()

        # Destroy toolbar and create a new one
        # (This has to be done, otherwise the toolbar won't function for a new plot)
        self.toolbar.destroy()
        self.toolbar = NavigationToolbar2Tk(self.plot_panel, self.toolbar_frame)
        self.toolbar.update()

        # Update title
        self.plotwindow.title(
            f"AutoGaitA Plot Panel {self.current_fig_index+1}/{len(self.figures)}"
        )


# %% what happens if we just hit run
if __name__ == "__main__":
    group_info_message = (
        "\n*************\nnot like this\n*************\n"
        + "You are trying to execute autogaita.group as a script, but that is not "
        + "possible.\nIf you prefer a non-GUI approach, please either: "
        + "\n1. Call this as a function, i.e. autogaita.group(folderinfo, cfg)"
        + "\n2. Use the dlc or simirun scripts in the batchrun_scripts folder"
    )
    print(group_info_message)
